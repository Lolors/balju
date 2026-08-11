"""거래명세서, 매입가, 운송비, 월별 매입 현황 관리 런처."""

from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import app_alias_fix as alias_fix

base_app = alias_fix.base_app

STATEMENTS_FILE = base_app.DATA / "purchase_statements.csv"
STATEMENT_ITEMS_FILE = base_app.DATA / "purchase_statement_items.csv"
PRICE_HISTORY_FILE = base_app.DATA / "price_history.csv"
MONTH_CLOSE_FILE = base_app.DATA / "purchase_month_close.csv"

STATEMENT_COLUMNS = [
    "명세서ID", "발주ID", "거래처명", "명세서번호", "명세서일자",
    "운송비", "운송비입력여부", "메모", "등록일시", "수정일시",
]
STATEMENT_ITEM_COLUMNS = [
    "명세서ID", "순번", "제품코드", "정식제품명", "규격", "단위", "발주수량",
    "입고수량", "매입단가", "상품금액", "출고단가", "가격적용여부",
]
PRICE_HISTORY_COLUMNS = [
    "가격ID", "명세서ID", "명세서일자", "제품코드", "정식제품명", "매입단가", "출고단가", "등록일시",
]
MONTH_CLOSE_COLUMNS = [
    "마감월", "상품매입금액", "운송비", "총매입금액", "거래명세서수", "등록일시", "메모",
]


def ensure_purchase_files():
    for path, columns in [
        (STATEMENTS_FILE, STATEMENT_COLUMNS),
        (STATEMENT_ITEMS_FILE, STATEMENT_ITEM_COLUMNS),
        (PRICE_HISTORY_FILE, PRICE_HISTORY_COLUMNS),
        (MONTH_CLOSE_FILE, MONTH_CLOSE_COLUMNS),
    ]:
        if not path.exists():
            path.write_text(",".join(columns) + "\n", encoding="utf-8-sig")


def read_table(path, columns):
    ensure_purchase_files()
    if path.exists():
        df = pd.read_csv(path, dtype=str).fillna("")
    else:
        df = pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns].fillna("")


def save_table(path, df, columns):
    clean = df.copy()
    for col in columns:
        if col not in clean.columns:
            clean[col] = ""
    clean = clean[columns].fillna("")
    clean.to_csv(path, index=False, encoding="utf-8-sig")


def load_purchase_data():
    return (
        read_table(STATEMENTS_FILE, STATEMENT_COLUMNS),
        read_table(STATEMENT_ITEMS_FILE, STATEMENT_ITEM_COLUMNS),
        read_table(PRICE_HISTORY_FILE, PRICE_HISTORY_COLUMNS),
        read_table(MONTH_CLOSE_FILE, MONTH_CLOSE_COLUMNS),
    )


def to_int(value):
    try:
        text = str(value).replace(",", "").strip()
        return int(float(text)) if text else 0
    except Exception:
        return 0


def money(value):
    return f"{to_int(value):,}"


def make_id(prefix, existing):
    today = datetime.now().strftime("%Y%m%d")
    base = f"{prefix}-{today}-"
    nums = []
    for value in existing:
        text = str(value)
        if text.startswith(base):
            tail = text.replace(base, "")
            if tail.isdigit():
                nums.append(int(tail))
    return f"{base}{(max(nums) + 1 if nums else 1):03d}"


def calc_sell_price(buy_price):
    return int(round(to_int(buy_price) * 1.3))


def normalize_statement_items(df):
    clean = df.copy().fillna("")
    for col in ["발주수량", "입고수량", "매입단가"]:
        clean[col] = clean[col].apply(to_int)
    clean["상품금액"] = clean["입고수량"] * clean["매입단가"]
    clean["출고단가"] = clean["매입단가"].apply(calc_sell_price)
    clean = clean[clean["입고수량"] > 0]
    return clean.reset_index(drop=True)


def get_received_by_order(statement_items, statements, order_id):
    if statement_items.empty or statements.empty:
        return pd.Series(dtype=int)
    ids = statements[statements["발주ID"] == order_id]["명세서ID"].tolist()
    related = statement_items[statement_items["명세서ID"].isin(ids)].copy()
    if related.empty:
        return pd.Series(dtype=int)
    related["입고수량"] = related["입고수량"].apply(to_int)
    return related.groupby("제품코드")["입고수량"].sum()


def page_statement_register(orders, order_items_saved):
    ensure_purchase_files()
    statements, statement_items, price_history, month_close = load_purchase_data()

    st.markdown("## 거래명세서 등록")
    st.caption("발주서 1건에 거래명세서를 여러 장 등록할 수 있습니다. 운송비는 월별 매입 합계에 항상 포함됩니다.")

    if orders.empty:
        st.info("등록된 발주서가 없습니다.")
        return

    order_options = orders.sort_values("발주일시", ascending=False)["발주ID"].tolist()
    selected_order = st.selectbox("연결할 발주서", order_options)
    order_header = orders[orders["발주ID"] == selected_order].iloc[0]
    order_rows = order_items_saved[order_items_saved["발주ID"] == selected_order].copy()

    if order_rows.empty:
        st.warning("이 발주서에는 품목이 없습니다.")
        return

    received = get_received_by_order(statement_items, statements, selected_order)
    order_rows["발주수량"] = order_rows["수량"].apply(to_int)
    order_rows["누적입고수량"] = order_rows["제품코드"].map(lambda code: int(received.get(code, 0)))
    order_rows["남은수량"] = order_rows["발주수량"] - order_rows["누적입고수량"]
    order_rows["이번입고수량"] = order_rows["남은수량"].apply(lambda x: max(0, int(x)))
    order_rows["매입단가"] = 0
    order_rows["가격적용여부"] = True

    with st.container(border=True):
        st.markdown("### 명세서 기본 정보")
        c1, c2, c3 = st.columns(3)
        with c1:
            statement_no = st.text_input("거래명세서 번호", value="")
            statement_date = st.date_input("거래명세서 일자", value=datetime.now().date())
        with c2:
            freight = st.number_input("운송비(배송비)", min_value=0, value=0, step=1000)
            freight_checked = st.checkbox("운송비 입력 완료", value=False, help="실제 운송비가 0원이어도 체크하면 누락으로 보지 않습니다.")
        with c3:
            memo = st.text_area("메모", height=88)

    st.markdown("### 품목별 매입가 입력")
    edit_cols = ["제품코드", "정식제품명", "규격", "단위", "발주수량", "누적입고수량", "남은수량", "이번입고수량", "매입단가", "가격적용여부"]
    edited = st.data_editor(
        order_rows[edit_cols],
        use_container_width=True,
        hide_index=True,
        disabled=["제품코드", "정식제품명", "규격", "단위", "발주수량", "누적입고수량", "남은수량"],
        column_config={
            "이번입고수량": st.column_config.NumberColumn("이번 입고수량", min_value=0, step=1),
            "매입단가": st.column_config.NumberColumn("매입단가", min_value=0, step=100),
            "가격적용여부": st.column_config.CheckboxColumn("현재 가격 적용"),
        },
        key=f"statement_items_{selected_order}",
    )

    preview = normalize_statement_items(edited.rename(columns={"이번입고수량": "입고수량"}))
    product_total = int(preview["상품금액"].sum()) if not preview.empty else 0
    total_amount = product_total + int(freight)

    a, b, c = st.columns(3)
    a.metric("상품 매입금액", f"{product_total:,}원")
    b.metric("운송비", f"{int(freight):,}원")
    c.metric("총 매입금액", f"{total_amount:,}원")

    if st.button("거래명세서 저장", type="primary", use_container_width=True):
        if not statement_no.strip():
            st.warning("거래명세서 번호를 입력하세요.")
            return
        if preview.empty:
            st.warning("입고수량이 1 이상인 품목을 입력하세요.")
            return
        if (preview["매입단가"] <= 0).any():
            st.warning("입고 품목의 매입단가를 입력하세요.")
            return

        sid = make_id("ST", statements["명세서ID"].tolist())
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_statement = pd.DataFrame([{
            "명세서ID": sid,
            "발주ID": selected_order,
            "거래처명": order_header["거래처명"],
            "명세서번호": statement_no.strip(),
            "명세서일자": statement_date.strftime("%Y-%m-%d"),
            "운송비": int(freight),
            "운송비입력여부": "Y" if freight_checked or int(freight) > 0 else "N",
            "메모": memo.strip(),
            "등록일시": now,
            "수정일시": now,
        }])

        item_rows = []
        price_rows = []
        existing_price_ids = price_history["가격ID"].tolist()
        for idx, row in preview.iterrows():
            item_rows.append({
                "명세서ID": sid,
                "순번": idx + 1,
                "제품코드": row["제품코드"],
                "정식제품명": row["정식제품명"],
                "규격": row["규격"],
                "단위": row["단위"],
                "발주수량": to_int(row["발주수량"]),
                "입고수량": to_int(row["입고수량"]),
                "매입단가": to_int(row["매입단가"]),
                "상품금액": to_int(row["상품금액"]),
                "출고단가": to_int(row["출고단가"]),
                "가격적용여부": "Y" if bool(row.get("가격적용여부", True)) else "N",
            })
            if bool(row.get("가격적용여부", True)):
                price_id = make_id("PR", existing_price_ids + [p.get("가격ID", "") for p in price_rows])
                price_rows.append({
                    "가격ID": price_id,
                    "명세서ID": sid,
                    "명세서일자": statement_date.strftime("%Y-%m-%d"),
                    "제품코드": row["제품코드"],
                    "정식제품명": row["정식제품명"],
                    "매입단가": to_int(row["매입단가"]),
                    "출고단가": to_int(row["출고단가"]),
                    "등록일시": now,
                })

        save_table(STATEMENTS_FILE, pd.concat([statements, new_statement], ignore_index=True), STATEMENT_COLUMNS)
        save_table(STATEMENT_ITEMS_FILE, pd.concat([statement_items, pd.DataFrame(item_rows)], ignore_index=True), STATEMENT_ITEM_COLUMNS)
        if price_rows:
            save_table(PRICE_HISTORY_FILE, pd.concat([price_history, pd.DataFrame(price_rows)], ignore_index=True), PRICE_HISTORY_COLUMNS)
        st.success(f"거래명세서를 저장했습니다: {sid}")
        st.rerun()


def statement_detail_frame(statements, statement_items):
    if statements.empty:
        return pd.DataFrame()
    items = statement_items.copy()
    for col in ["입고수량", "매입단가", "상품금액", "출고단가"]:
        if col in items.columns:
            items[col] = items[col].apply(to_int)
    item_sum = items.groupby("명세서ID", as_index=False).agg(상품매입금액=("상품금액", "sum"), 매입품목수=("제품코드", "count"), 총입고수량=("입고수량", "sum")) if not items.empty else pd.DataFrame(columns=["명세서ID", "상품매입금액", "매입품목수", "총입고수량"])
    result = statements.merge(item_sum, on="명세서ID", how="left").fillna("")
    for col in ["상품매입금액", "매입품목수", "총입고수량", "운송비"]:
        result[col] = result[col].apply(to_int)
    result["총매입금액"] = result["상품매입금액"] + result["운송비"]
    result["운송비상태"] = result["운송비입력여부"].map(lambda v: "입력완료" if str(v) == "Y" else "미입력")
    return result


def page_statement_list():
    statements, statement_items, _, _ = load_purchase_data()
    st.markdown("## 거래명세서 내역")
    if statements.empty:
        st.info("등록된 거래명세서가 없습니다.")
        return
    detail = statement_detail_frame(statements, statement_items).sort_values("명세서일자", ascending=False)
    show = detail[["명세서ID", "발주ID", "거래처명", "명세서번호", "명세서일자", "상품매입금액", "운송비", "총매입금액", "운송비상태"]].copy()
    for col in ["상품매입금액", "운송비", "총매입금액"]:
        show[col] = show[col].apply(lambda v: f"{to_int(v):,}")
    st.dataframe(show, use_container_width=True, hide_index=True)

    missing = detail[detail["운송비입력여부"] != "Y"]
    if not missing.empty:
        st.warning(f"운송비 미입력 거래명세서 {len(missing)}건이 있습니다.")
        with st.expander("운송비 미입력 거래명세서 보기", expanded=True):
            miss_show = missing[["명세서ID", "발주ID", "거래처명", "명세서번호", "명세서일자", "상품매입금액", "운송비"]].copy()
            for col in ["상품매입금액", "운송비"]:
                miss_show[col] = miss_show[col].apply(lambda v: f"{to_int(v):,}")
            st.dataframe(miss_show, use_container_width=True, hide_index=True)


def filter_month(df, date_col, year, month):
    if df.empty:
        return df.copy()
    temp = df.copy()
    temp[date_col] = pd.to_datetime(temp[date_col], errors="coerce")
    return temp[(temp[date_col].dt.year == year) & (temp[date_col].dt.month == month)].copy()


def monthly_excel_bytes(month_key, summary, vendor_group, product_group, detail):
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name, df):
        ws = wb.create_sheet(name)
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F1F5F9")
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row, 1):
                ws.cell(r, c, value)
        for c in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

    add_sheet("월요약", pd.DataFrame([summary]))
    add_sheet("거래처별", vendor_group)
    add_sheet("제품별", product_group)
    add_sheet("명세서상세", detail)
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def page_monthly_purchase():
    statements, statement_items, price_history, month_close = load_purchase_data()
    st.markdown("## 월별 매입 현황")
    st.caption("월별 매입 합계는 거래명세서 일자 기준이며, 운송비는 항상 총매입금액에 포함됩니다.")

    today = datetime.now()
    c1, c2 = st.columns(2)
    year = c1.number_input("연도", min_value=2020, max_value=2100, value=today.year, step=1)
    month = c2.number_input("월", min_value=1, max_value=12, value=today.month, step=1)
    month_key = f"{int(year):04d}-{int(month):02d}"

    month_statements = filter_month(statements, "명세서일자", int(year), int(month))
    if month_statements.empty:
        st.info("선택한 월의 거래명세서가 없습니다.")
        return

    month_items = statement_items[statement_items["명세서ID"].isin(month_statements["명세서ID"].tolist())].copy()
    detail = statement_detail_frame(month_statements, month_items)
    product_amount = int(detail["상품매입금액"].sum())
    freight_total = int(detail["운송비"].sum())
    total_amount = product_amount + freight_total
    missing_freight = detail[detail["운송비입력여부"] != "Y"]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("상품 매입금액", f"{product_amount:,}원")
    m2.metric("운송비", f"{freight_total:,}원")
    m3.metric("총 매입금액", f"{total_amount:,}원")
    m4.metric("운송비 미입력", f"{len(missing_freight):,}건")

    if not missing_freight.empty:
        st.warning(f"운송비가 입력되지 않은 거래명세서가 {len(missing_freight)}건 있습니다.")
        with st.expander("운송비 미입력 거래명세서 보기", expanded=True):
            st.dataframe(missing_freight[["명세서ID", "발주ID", "거래처명", "명세서번호", "명세서일자", "상품매입금액", "운송비"]], use_container_width=True, hide_index=True)

    vendor_group = detail.groupby("거래처명", as_index=False).agg(
        거래명세서수=("명세서ID", "count"),
        상품매입금액=("상품매입금액", "sum"),
        운송비=("운송비", "sum"),
        총매입금액=("총매입금액", "sum"),
    )
    st.markdown("### 거래처별 집계")
    st.dataframe(vendor_group, use_container_width=True, hide_index=True)

    if not month_items.empty:
        month_items["입고수량"] = month_items["입고수량"].apply(to_int)
        month_items["상품금액"] = month_items["상품금액"].apply(to_int)
        month_items["매입단가"] = month_items["매입단가"].apply(to_int)
        product_group = month_items.groupby(["제품코드", "정식제품명", "규격", "단위"], as_index=False).agg(
            총입고수량=("입고수량", "sum"),
            월매입금액=("상품금액", "sum"),
        )
        latest_price = price_history.copy()
        if not latest_price.empty:
            latest_price["명세서일자"] = pd.to_datetime(latest_price["명세서일자"], errors="coerce")
            latest_price = latest_price.sort_values(["제품코드", "명세서일자", "등록일시"]).groupby("제품코드", as_index=False).tail(1)
            latest_price = latest_price[["제품코드", "매입단가", "출고단가"]].rename(columns={"매입단가": "최근매입단가", "출고단가": "현재출고가"})
            product_group = product_group.merge(latest_price, on="제품코드", how="left")
    else:
        product_group = pd.DataFrame(columns=["제품코드", "정식제품명", "규격", "단위", "총입고수량", "월매입금액", "최근매입단가", "현재출고가"])

    st.markdown("### 제품별 집계")
    st.dataframe(product_group, use_container_width=True, hide_index=True)

    detail_export = detail[["명세서ID", "발주ID", "거래처명", "명세서번호", "명세서일자", "상품매입금액", "운송비", "총매입금액", "운송비상태"]].copy()
    summary = {
        "마감월": month_key,
        "상품매입금액": product_amount,
        "운송비": freight_total,
        "총매입금액": total_amount,
        "거래명세서수": len(detail),
        "운송비미입력건수": len(missing_freight),
    }
    st.download_button(
        "월별 매입 현황 엑셀 다운로드",
        data=monthly_excel_bytes(month_key, summary, vendor_group, product_group, detail_export),
        file_name=f"월별매입현황_{month_key}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    with st.container(border=True):
        st.markdown("### 월마감")
        if not missing_freight.empty:
            st.warning("운송비 미입력 거래명세서가 남아 있습니다. 실제로 0원이라면 해당 명세서에서 운송비 입력 완료 처리 후 마감하는 것을 권장합니다.")
        memo = st.text_input("마감 메모", value="")
        if st.button(f"{month_key} 매입 마감 저장", type="primary", use_container_width=True):
            close_row = pd.DataFrame([{
                "마감월": month_key,
                "상품매입금액": product_amount,
                "운송비": freight_total,
                "총매입금액": total_amount,
                "거래명세서수": len(detail),
                "등록일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "메모": memo,
            }])
            new_close = month_close[month_close["마감월"] != month_key]
            save_table(MONTH_CLOSE_FILE, pd.concat([new_close, close_row], ignore_index=True), MONTH_CLOSE_COLUMNS)
            st.success("월마감 내역을 저장했습니다.")
            st.rerun()


def render_sidebar():
    st.sidebar.markdown('<div class="sidebar-title">발주관리 시스템</div>', unsafe_allow_html=True)

    def menu(name):
        active = st.session_state.current_page == name
        if st.sidebar.button(name, key=f"purchase_menu_{name}", use_container_width=True, type="primary" if active else "secondary"):
            st.session_state.current_page = name
            st.rerun()

    menu("발주 작성")
    menu("임시저장 목록")
    menu("발주서 목록")
    menu("최근 발주 내역")
    st.sidebar.markdown("---")
    st.sidebar.markdown("**매입 관리**")
    menu("거래명세서 등록")
    menu("거래명세서 내역")
    menu("월별 매입 현황")
    st.sidebar.markdown("---")
    st.sidebar.markdown("**기초 관리**")
    menu("거래처 관리")
    menu("제품 관리")
    menu("별칭 관리")
    return st.session_state.current_page


def main():
    ensure_purchase_files()
    base_app.inject_css()
    base_app.init_state()
    vendors, products, aliases, drafts, draft_items, orders, order_items_saved = base_app.load_data()
    page = render_sidebar()

    if page == "발주 작성":
        base_app.page_write(vendors, products, aliases, drafts, draft_items, orders, order_items_saved)
    elif page == "임시저장 목록":
        base_app.page_drafts(drafts, draft_items)
    elif page == "발주서 목록":
        base_app.page_orders(vendors, orders, order_items_saved)
    elif page == "최근 발주 내역":
        base_app.page_recent_orders(orders)
    elif page == "거래명세서 등록":
        page_statement_register(orders, order_items_saved)
    elif page == "거래명세서 내역":
        page_statement_list()
    elif page == "월별 매입 현황":
        page_monthly_purchase()
    elif page == "거래처 관리":
        base_app.page_vendor_manage(vendors)
    elif page == "제품 관리":
        base_app.page_product_manage(products)
    elif page == "별칭 관리":
        base_app.page_alias_manage(vendors, products, aliases)
    else:
        base_app.page_placeholder(page)


if __name__ == "__main__":
    main()
