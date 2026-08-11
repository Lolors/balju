"""
PDF/PNG 다운로드 흐름 개선 런처.

기존 app.py를 그대로 불러온 뒤 미리보기 내보내기 함수와 일부 화면만 교체합니다.
- PDF/PNG 버튼을 '생성'과 '저장'으로 분리
- 생성 성공 메시지와 저장 폴더 표시
- 발주 내용이 바뀌면 이전 PDF/PNG 다운로드 상태 초기화
- 생성 후 st.rerun() 없이 같은 화면에서 바로 저장 버튼 표시
- 엑셀/PDF/PNG 저장 안내 메시지는 3초 뒤 자동으로 숨김
- 발주서 미리보기 품목 표는 No. / 제품명 / 규격 / 수량 / 단위로 표시
- 요청사항이 비어 있으면 발주서 미리보기에서 요청사항 박스를 숨김
- 제품 관리는 엑셀 양식 다운로드/업로드를 지원하고, 수정 저장 시 체크 행이 사라지지 않게 처리
"""

import hashlib
import json
import re
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

import app as base_app


EXPORT_MESSAGE_SECONDS = 3
PRODUCT_COLUMNS = ["제품코드", "정식제품명", "규격", "단위"]
ORIGINAL_CREATE_EXCEL = base_app.create_excel


def hide_empty_request_box(html, request_note):
    """요청사항이 없으면 요청사항 박스를 표시하지 않습니다."""
    if str(request_note or "").strip():
        return html
    return re.sub(
        r"\s*<div class=\"request\">\s*<b>요청사항</b><br>\s*-\s*</div>\s*",
        "\n",
        html,
        flags=re.DOTALL,
    )


def prepare_preview_template(template):
    """미리보기 표에서 제품코드 컬럼을 제거하고 수량/단위 순서로 맞춥니다."""
    template = template.replace(
        "<th>No.</th>\n                <th>제품코드</th>\n                <th>제품명</th>\n                <th>규격</th>\n                <th>단위</th>\n                <th>수량</th>",
        "<th>No.</th>\n                <th>제품명</th>\n                <th>규격</th>\n                <th>수량</th>\n                <th>단위</th>",
    )
    template = template.replace(
        "<th>No.</th>\n                <th>제품코드</th>\n                <th>제품명</th>\n                <th>규격</th>\n                <th>수량</th>\n                <th>단위</th>",
        "<th>No.</th>\n                <th>제품명</th>\n                <th>규격</th>\n                <th>수량</th>\n                <th>단위</th>",
    )
    template = template.replace(
        ".item-table th:nth-child(2),\n.item-table td:nth-child(2) {\n    width: 92px;\n}\n\n.item-table th:nth-child(3),\n.item-table td:nth-child(3) {\n    width: 230px;\n}",
        ".item-table th:nth-child(2),\n.item-table td:nth-child(2) {\n    width: 280px;\n}\n\n.item-table th:nth-child(3),\n.item-table td:nth-child(3) {\n    width: 120px;\n}",
    )
    return template


def render_order_html(vendor, order_items, request_note, order_id=None, order_date=None):
    """발주서 미리보기 HTML을 제품코드 없이 직접 렌더링합니다."""
    template = base_app.TEMPLATE_FILE.read_text(encoding="utf-8") if base_app.TEMPLATE_FILE.exists() else base_app.DEFAULT_TEMPLATE
    template = prepare_preview_template(template)
    logo_b64 = base_app.get_logo_base64()

    order_id = order_id or f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    order_date = order_date or datetime.now().strftime("%Y-%m-%d")

    rows_html = ""
    for idx, item in enumerate(order_items, 1):
        rows_html += f"""
        <tr>
            <td>{idx}</td>
            <td class="left">{item.get("정식제품명", "")}</td>
            <td>{item.get("규격", "")}</td>
            <td>{base_app.fmt_int(item.get("수량", 0))}</td>
            <td>{item.get("단위", "")}</td>
        </tr>
        """

    if not rows_html:
        rows_html = '<tr><td colspan="5" class="empty">발주 품목이 없습니다.</td></tr>'

    total_count, total_qty = base_app.calc_totals(order_items)

    if logo_b64:
        template = template.replace("{% if_logo %}", "").replace("{% else_logo %}", "<!--").replace("{% endif_logo %}", "-->")
    else:
        template = template.replace("{% if_logo %}", "<!--").replace("{% else_logo %}", "-->").replace("{% endif_logo %}", "")

    html = (
        template
        .replace("{{LOGO_BASE64}}", logo_b64)
        .replace("{{ORDER_ID}}", order_id)
        .replace("{{ORDER_DATE}}", order_date)
        .replace("{{COMPANY_NAME}}", base_app.COMPANY["상호"])
        .replace("{{COMPANY_OWNER}}", base_app.COMPANY["대표"])
        .replace("{{COMPANY_ADDRESS}}", base_app.COMPANY["주소"])
        .replace("{{COMPANY_PHONE}}", base_app.COMPANY["연락처"])
        .replace("{{COMPANY_FAX}}", base_app.COMPANY["팩스"])
        .replace("{{COMPANY_REGNO}}", base_app.COMPANY["등록번호"])
        .replace("{{VENDOR_NAME}}", str(vendor.get("거래처명", "")))
        .replace("{{VENDOR_ADDRESS}}", str(vendor.get("배송지", "")))
        .replace("{{VENDOR_PHONE}}", str(vendor.get("연락처", "")))
        .replace("{{ITEM_ROWS}}", rows_html)
        .replace("{{TOTAL_COUNT}}", base_app.fmt_int(total_count))
        .replace("{{TOTAL_QTY}}", base_app.fmt_int(total_qty))
        .replace("{{REQUEST_NOTE}}", request_note or "")
    )
    return hide_empty_request_box(html, request_note)


def create_excel(vendor, order_items, request_note, order_date=None):
    """엑셀 저장 파일은 기존처럼 제품코드를 유지하되 수량/단위 순서로 맞춥니다."""
    path = ORIGINAL_CREATE_EXCEL(vendor, order_items, request_note, order_date)
    wb = load_workbook(path)
    ws = wb.active

    start_row = 14
    ws.cell(start_row, 5, "수량")
    ws.cell(start_row, 6, "단위")

    for idx, item in enumerate(order_items, 1):
        row_num = start_row + idx
        ws.cell(row_num, 5, base_app.safe_int(item.get("수량", 0)))
        ws.cell(row_num, 6, item.get("단위", ""))

    if not str(request_note or "").strip():
        total_row = start_row + len(order_items) + 2
        ws.cell(total_row + 3, 1, "")
        ws.cell(total_row + 4, 1, "")

    wb.save(path)
    return path


def normalize_product_table(df):
    """제품 엑셀/편집 데이터를 저장 가능한 형태로 정리합니다."""
    df = df.copy()
    for col in PRODUCT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[PRODUCT_COLUMNS].fillna("")
    for col in PRODUCT_COLUMNS:
        df[col] = df[col].astype(str).str.strip()

    # 완전히 빈 줄과 제품명이 없는 줄은 저장하지 않습니다.
    df = df[df[PRODUCT_COLUMNS].apply(lambda row: any(str(v).strip() for v in row), axis=1)]
    df = df[df["정식제품명"].str.strip() != ""]
    return df.reset_index(drop=True)


def product_excel_bytes(df=None):
    """제품 업로드용 엑셀 양식을 만듭니다."""
    if df is None or df.empty:
        df = pd.DataFrame(columns=PRODUCT_COLUMNS)
    else:
        df = normalize_product_table(df)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="제품목록")
        ws = writer.book["제품목록"]
        widths = {"A": 18, "B": 34, "C": 18, "D": 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    output.seek(0)
    return output.getvalue()


def read_product_upload(uploaded_file):
    """업로드된 제품 엑셀/CSV를 읽습니다."""
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(uploaded_file, dtype=str).fillna("")
    else:
        raw = pd.read_excel(uploaded_file, dtype=str).fillna("")

    missing = [col for col in PRODUCT_COLUMNS if col not in raw.columns]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))

    clean = normalize_product_table(raw)
    duplicate_codes = clean[clean["제품코드"] != ""]["제품코드"]
    if duplicate_codes.duplicated().any():
        duplicated = duplicate_codes[duplicate_codes.duplicated()].drop_duplicates().tolist()
        raise ValueError("중복 제품코드가 있습니다: " + ", ".join(duplicated[:10]))

    return clean


def page_product_manage(products):
    st.markdown("## 제품 관리")
    st.caption("제품코드, 정식제품명, 규격, 단위를 관리합니다. 단가/금액은 사용하지 않습니다.")

    with st.container(border=True):
        st.markdown("### 신규 제품 추가")

        c1, c2, c3 = st.columns(3)
        with c1:
            new_code = st.text_input("제품코드", value="", placeholder="업체 고유코드가 있으면 입력")
            new_name = st.text_input("정식제품명")
        with c2:
            new_spec = st.text_input("규격")
            new_unit = st.text_input("단위", value="EA")
        if st.button("제품 추가", type="primary", use_container_width=True):
            if not new_name.strip():
                st.warning("정식제품명을 입력하세요.")
            elif new_code.strip() and new_code.strip() in products["제품코드"].tolist():
                st.warning("이미 존재하는 제품코드입니다.")
            else:
                new_row = pd.DataFrame([{
                    "제품코드": new_code.strip(),
                    "정식제품명": new_name.strip(),
                    "규격": new_spec.strip(),
                    "단위": new_unit.strip(),
                }])
                products = pd.concat([products, new_row], ignore_index=True)
                base_app.save_products(products)
                st.success("제품을 추가했습니다.")
                st.rerun()

    with st.container(border=True):
        st.markdown("### 엑셀로 제품 목록 일괄 등록/수정")
        st.caption("양식의 컬럼명은 반드시 제품코드 / 정식제품명 / 규격 / 단위 순서로 유지하세요. 업로드 저장 시 현재 제품 목록이 업로드 파일 기준으로 교체됩니다.")

        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "제품 업로드 양식 다운로드",
                data=product_excel_bytes(),
                file_name="제품목록_업로드양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "현재 제품 목록 다운로드",
                data=product_excel_bytes(products),
                file_name="현재_제품목록.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        uploaded = st.file_uploader("제품 목록 엑셀 업로드", type=["xlsx", "xls", "csv"], key="product_bulk_upload")
        if uploaded is not None:
            try:
                uploaded_products = read_product_upload(uploaded)
                st.success(f"업로드 파일에서 {len(uploaded_products)}개 제품을 읽었습니다.")
                st.dataframe(uploaded_products, use_container_width=True, hide_index=True)

                if st.button("업로드 파일로 제품 목록 저장", type="primary", use_container_width=True):
                    base_app.save_products(uploaded_products)
                    st.success("업로드한 제품 목록으로 저장했습니다.")
                    st.rerun()
            except Exception as e:
                st.error(f"제품 목록 업로드 실패: {e}")

    st.markdown("### 제품 목록 수정")

    if products.empty:
        st.info("등록된 제품이 없습니다.")
        return

    edit_df = products.copy()
    edit_df["삭제"] = False
    edit_df = edit_df[["삭제", "제품코드", "정식제품명", "규격", "단위"]]

    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "삭제": st.column_config.CheckboxColumn("삭제"),
        },
        key="product_editor",
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("제품 수정 저장", use_container_width=True):
            # 수정 저장은 체크박스를 무시하고 현재 편집 내용을 그대로 저장합니다.
            clean = normalize_product_table(edited.drop(columns=["삭제"]))
            base_app.save_products(clean)
            st.success("제품 정보를 저장했습니다.")
            st.rerun()

    with c2:
        if st.button("선택 제품 삭제", use_container_width=True):
            clean = edited[edited["삭제"] != True].drop(columns=["삭제"])
            clean = normalize_product_table(clean)
            base_app.save_products(clean)
            st.success("선택한 제품을 삭제했습니다.")
            st.rerun()


base_app.render_order_html = render_order_html
base_app.create_excel = create_excel
base_app.page_product_manage = page_product_manage


def make_preview_signature(vendor, order_items, request_note, order_date):
    vendor_payload = {
        "거래처명": str(vendor.get("거래처명", "")),
        "담당자": str(vendor.get("담당자", "")),
        "연락처": str(vendor.get("연락처", "")),
        "배송지": str(vendor.get("배송지", "")),
    }
    payload = {
        "vendor": vendor_payload,
        "items": order_items,
        "request_note": request_note or "",
        "order_date": str(order_date or ""),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def set_export_message(message):
    st.session_state["preview_export_message"] = message
    st.session_state["preview_export_message_time"] = time.time()


def reset_preview_export_state():
    for key in [
        "preview_pdf_path",
        "preview_png_path",
        "preview_capture_error",
        "preview_export_message",
        "preview_export_message_time",
    ]:
        st.session_state.pop(key, None)


def show_temporary_export_message():
    capture_error = st.session_state.get("preview_capture_error")
    export_message = st.session_state.get("preview_export_message")
    message_time = st.session_state.get("preview_export_message_time")

    if capture_error:
        st.error(
            "PDF/PNG 캡쳐 생성에 실패했습니다. 앱을 실행한 Python 환경에서 Playwright와 Chromium이 설치되어 있는지 확인하세요.\n\n"
            "확인 명령:\n"
            "python -c \"from playwright.sync_api import sync_playwright; p=sync_playwright().start(); b=p.chromium.launch(headless=True); print('OK'); b.close(); p.stop()\"\n\n"
            f"오류내용: {capture_error}"
        )
        return

    if not export_message or not message_time:
        return

    elapsed = time.time() - float(message_time)
    if elapsed >= EXPORT_MESSAGE_SECONDS:
        st.session_state.pop("preview_export_message", None)
        st.session_state.pop("preview_export_message_time", None)
        return

    placeholder = st.empty()
    with placeholder.container():
        st.success(export_message)
        st.caption(f"저장 폴더: {base_app.PDF_OUTPUT}")

    time.sleep(max(0, EXPORT_MESSAGE_SECONDS - elapsed))
    placeholder.empty()
    st.session_state.pop("preview_export_message", None)
    st.session_state.pop("preview_export_message_time", None)


def render_purchase_preview(vendor, order_items, request_note, order_date=None):
    preview_signature = make_preview_signature(vendor, order_items, request_note, order_date)
    if st.session_state.get("preview_export_signature") != preview_signature:
        reset_preview_export_state()
        st.session_state["preview_export_signature"] = preview_signature

    top_left, top_excel, top_pdf, top_png = st.columns([2.25, 0.75, 0.85, 0.85])

    with top_left:
        st.markdown('<div class="preview-title">발주서 미리보기</div>', unsafe_allow_html=True)

    with top_excel:
        excel_path = base_app.create_excel(vendor, order_items, request_note, order_date)
        st.download_button(
            "엑셀 저장",
            excel_path.read_bytes(),
            file_name=excel_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"preview_xls_{len(order_items)}_{order_date}",
            help="엑셀 내려받기",
            on_click=set_export_message,
            args=(f"엑셀 저장 준비 완료: {excel_path}",),
        )

    with top_pdf:
        if st.button("PDF 생성", use_container_width=True, key="preview_pdf_generate", help="PDF 파일 생성"):
            try:
                with st.spinner("PDF 생성 중..."):
                    pdf_path = base_app.create_preview_pdf(vendor, order_items, request_note, order_date)
                st.session_state["preview_pdf_path"] = str(pdf_path)
                set_export_message(f"PDF 생성 완료: {pdf_path}")
                st.session_state.pop("preview_capture_error", None)
            except Exception as e:
                st.session_state["preview_capture_error"] = str(e)
                st.session_state.pop("preview_pdf_path", None)

        pdf_saved_path = st.session_state.get("preview_pdf_path")
        if pdf_saved_path and Path(pdf_saved_path).exists():
            pdf_path = Path(pdf_saved_path)
            st.download_button(
                "PDF 저장",
                pdf_path.read_bytes(),
                file_name=pdf_path.name,
                mime="application/pdf",
                use_container_width=True,
                key=f"preview_pdf_download_{pdf_path.name}",
                help="생성된 PDF 내려받기",
                on_click=set_export_message,
                args=(f"PDF 저장 준비 완료: {pdf_path}",),
            )

    with top_png:
        if st.button("PNG 생성", use_container_width=True, key="preview_png_generate", help="PNG 파일 생성"):
            try:
                with st.spinner("PNG 생성 중..."):
                    png_path = base_app.create_preview_image(vendor, order_items, request_note, order_date)
                st.session_state["preview_png_path"] = str(png_path)
                set_export_message(f"PNG 생성 완료: {png_path}")
                st.session_state.pop("preview_capture_error", None)
            except Exception as e:
                st.session_state["preview_capture_error"] = str(e)
                st.session_state.pop("preview_png_path", None)

        png_saved_path = st.session_state.get("preview_png_path")
        if png_saved_path and Path(png_saved_path).exists():
            png_path = Path(png_saved_path)
            st.download_button(
                "PNG 저장",
                png_path.read_bytes(),
                file_name=png_path.name,
                mime="image/png",
                use_container_width=True,
                key=f"preview_png_download_{png_path.name}",
                help="생성된 PNG 내려받기",
                on_click=set_export_message,
                args=(f"PNG 저장 준비 완료: {png_path}",),
            )

    show_temporary_export_message()

    html = base_app.render_order_html(vendor, order_items, request_note, order_date=order_date)
    components.html(html, height=790, scrolling=True)


base_app.render_purchase_preview = render_purchase_preview

if __name__ == "__main__":
    base_app.main()
