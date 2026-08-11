"""제품유형/제품코드/제품명/규격/포장단위 스키마와 발주서 표시를 적용하는 실행 런처."""

import re
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import app_statement_detail as current

purchase = current.purchase
base_app = current.base_app
st = base_app.st

EXTERNAL_PRODUCT_COLUMNS = ["제품유형", "제품코드", "제품명", "규격", "포장단위"]
ORIGINAL_LOAD_DATA = base_app.load_data
ORIGINAL_CREATE_EXCEL = base_app.create_excel


def normalize_external_products(df):
    """신규/기존 제품 파일을 새 제품 스키마로 정리합니다."""
    clean = df.copy().fillna("")

    # 기존 컬럼과 새 컬럼을 모두 허용합니다.
    if "제품명" not in clean.columns:
        clean["제품명"] = clean.get("정식제품명", "")
    if "포장단위" not in clean.columns:
        clean["포장단위"] = clean.get("단위", "")
    if "제품유형" not in clean.columns:
        clean["제품유형"] = ""

    for col in EXTERNAL_PRODUCT_COLUMNS:
        if col not in clean.columns:
            clean[col] = ""
        clean[col] = clean[col].astype(str).str.strip()

    clean = clean[EXTERNAL_PRODUCT_COLUMNS]
    clean = clean[clean.apply(lambda row: any(str(v).strip() for v in row), axis=1)]
    clean = clean[clean["제품명"].str.strip() != ""]
    return clean.reset_index(drop=True)


def products_for_app(df):
    """앱의 기존 검색/별칭 기능이 사용할 호환 컬럼을 함께 제공합니다."""
    external = normalize_external_products(df)
    internal = external.copy()
    internal["정식제품명"] = internal["제품명"]
    internal["단위"] = internal["포장단위"]
    return internal


def read_products_file():
    if base_app.PRODUCTS_FILE.exists():
        raw = pd.read_csv(base_app.PRODUCTS_FILE, dtype=str).fillna("")
    else:
        raw = pd.DataFrame(columns=EXTERNAL_PRODUCT_COLUMNS)
    return products_for_app(raw)


def load_data_with_product_schema():
    vendors, _, aliases, drafts, draft_items, orders, order_items = ORIGINAL_LOAD_DATA()
    products = read_products_file()
    return vendors, products, aliases, drafts, draft_items, orders, order_items


def save_products_with_schema(df):
    external = normalize_external_products(df)
    external.to_csv(base_app.PRODUCTS_FILE, index=False, encoding="utf-8-sig")


def product_excel_bytes(df=None):
    external = normalize_external_products(df) if df is not None else pd.DataFrame(columns=EXTERNAL_PRODUCT_COLUMNS)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        external.to_excel(writer, index=False, sheet_name="제품목록")
        ws = writer.book["제품목록"]
        widths = {"A": 16, "B": 18, "C": 36, "D": 20, "E": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    output.seek(0)
    return output.getvalue()


def read_product_upload(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(uploaded_file, dtype=str).fillna("")
    else:
        raw = pd.read_excel(uploaded_file, dtype=str).fillna("")

    missing = [col for col in EXTERNAL_PRODUCT_COLUMNS if col not in raw.columns]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))

    clean = normalize_external_products(raw)
    codes = clean.loc[clean["제품코드"] != "", "제품코드"]
    if codes.duplicated().any():
        duplicated = codes[codes.duplicated()].drop_duplicates().tolist()
        raise ValueError("중복 제품코드가 있습니다: " + ", ".join(duplicated[:10]))
    return clean


def page_product_manage_with_schema(products):
    st.markdown("## 제품 관리")
    st.caption("제품유형, 제품코드, 제품명, 규격, 포장단위를 관리합니다.")

    external_products = normalize_external_products(products)

    with st.container(border=True):
        st.markdown("### 신규 제품 추가")
        c1, c2, c3 = st.columns(3)
        with c1:
            new_type = st.text_input("제품유형")
            new_code = st.text_input("제품코드", value="")
        with c2:
            new_name = st.text_input("제품명")
            new_spec = st.text_input("규격")
        with c3:
            new_pack = st.text_input("포장단위")

        if st.button("제품 추가", type="primary", use_container_width=True):
            if not new_name.strip():
                st.warning("제품명을 입력하세요.")
            elif new_code.strip() and new_code.strip() in external_products["제품코드"].tolist():
                st.warning("이미 존재하는 제품코드입니다.")
            else:
                new_row = pd.DataFrame([{
                    "제품유형": new_type.strip(),
                    "제품코드": new_code.strip(),
                    "제품명": new_name.strip(),
                    "규격": new_spec.strip(),
                    "포장단위": new_pack.strip(),
                }])
                save_products_with_schema(pd.concat([external_products, new_row], ignore_index=True))
                st.success("제품을 추가했습니다.")
                st.rerun()

    with st.container(border=True):
        st.markdown("### 엑셀로 제품 목록 일괄 등록/수정")
        st.caption("파일 컬럼은 제품유형 / 제품코드 / 제품명 / 규격 / 포장단위 순서로 준비하세요. 저장하면 현재 제품 목록이 업로드 파일 기준으로 교체됩니다.")

        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "제품 업로드 양식 다운로드",
                data=product_excel_bytes(),
                file_name="제품목록_업로드양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "현재 제품 목록 다운로드",
                data=product_excel_bytes(external_products),
                file_name="현재_제품목록.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        uploaded = st.file_uploader("제품 목록 엑셀 업로드", type=["xlsx", "xls", "csv"], key="product_bulk_upload_v2")
        if uploaded is not None:
            try:
                uploaded_products = read_product_upload(uploaded)
                st.success(f"업로드 파일에서 {len(uploaded_products)}개 제품을 읽었습니다.")
                st.dataframe(uploaded_products, use_container_width=True, hide_index=True)
                if st.button("업로드 파일로 제품 목록 저장", type="primary", use_container_width=True):
                    save_products_with_schema(uploaded_products)
                    st.success("업로드한 제품 목록으로 저장했습니다.")
                    st.rerun()
            except Exception as exc:
                st.error(f"제품 목록 업로드 실패: {exc}")

    st.markdown("### 제품 목록 수정")
    if external_products.empty:
        st.info("등록된 제품이 없습니다.")
        return

    edit_df = external_products.copy()
    edit_df["삭제"] = False
    edit_df = edit_df[["삭제"] + EXTERNAL_PRODUCT_COLUMNS]
    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={"삭제": st.column_config.CheckboxColumn("삭제")},
        key="product_editor_v2",
    )

    c1, c2 = st.columns(2)
    with c1:
        if st.button("제품 수정 저장", use_container_width=True):
            save_products_with_schema(edited.drop(columns=["삭제"]))
            st.success("제품 정보를 저장했습니다.")
            st.rerun()
    with c2:
        if st.button("선택 제품 삭제", use_container_width=True):
            remaining = edited[edited["삭제"] != True].drop(columns=["삭제"])
            save_products_with_schema(remaining)
            st.success("선택한 제품을 삭제했습니다.")
            st.rerun()


def prepare_template_for_packaging(template):
    """발주서 표 머리글을 제품명/규격/포장단위/수량 순서로 맞춥니다."""
    pattern = r"<thead>\s*<tr>.*?</tr>\s*</thead>"
    replacement = """<thead>
            <tr>
                <th>No.</th>
                <th>제품명</th>
                <th>규격</th>
                <th>포장단위</th>
                <th>수량</th>
            </tr>
        </thead>"""
    return re.sub(pattern, replacement, template, count=1, flags=re.DOTALL)


def render_order_html_with_packaging(vendor, order_items, request_note, order_id=None, order_date=None):
    template = base_app.TEMPLATE_FILE.read_text(encoding="utf-8") if base_app.TEMPLATE_FILE.exists() else base_app.DEFAULT_TEMPLATE
    template = prepare_template_for_packaging(template)
    logo_b64 = base_app.get_logo_base64()

    from datetime import datetime
    order_id = order_id or f"PO-{datetime.now().strftime('%y%m%d-%H%M%S')}"
    order_date = order_date or datetime.now().strftime("%Y-%m-%d")

    rows_html = ""
    for idx, item in enumerate(order_items, 1):
        pack = item.get("포장단위", item.get("단위", ""))
        rows_html += f"""
        <tr>
            <td>{idx}</td>
            <td class="left">{item.get("제품명", item.get("정식제품명", ""))}</td>
            <td>{item.get("규격", "")}</td>
            <td>{pack}</td>
            <td>{base_app.fmt_int(item.get("수량", 0))}</td>
        </tr>
        """

    if not rows_html:
        rows_html = '<tr><td colspan="5" class="empty">발주 품목이 없습니다.</td></tr>'

    total_count, total_qty = base_app.calc_totals(order_items)
    if logo_b64:
        template = template.replace("{% if_logo %}", "").replace("{% else_logo %}", "<!--").replace("{% endif_logo %}", "-->")
    else:
        template = template.replace("{% if_logo %}", "<!--").replace("{% else_logo %}", "-->").replace("{% endif_logo %}", "")

    html = (
        template
        .replace("{{LOGO_BASE64}}", logo_b64)
        .replace("{{ORDER_ID}}", order_id)
        .replace("{{ORDER_DATE}}", order_date)
        .replace("{{COMPANY_NAME}}", base_app.COMPANY["상호"])
        .replace("{{COMPANY_OWNER}}", base_app.COMPANY["대표"])
        .replace("{{COMPANY_ADDRESS}}", base_app.COMPANY["주소"])
        .replace("{{COMPANY_PHONE}}", base_app.COMPANY["연락처"])
        .replace("{{COMPANY_FAX}}", base_app.COMPANY["팩스"])
        .replace("{{COMPANY_REGNO}}", base_app.COMPANY["등록번호"])
        .replace("{{VENDOR_NAME}}", str(vendor.get("거래처명", "")))
        .replace("{{VENDOR_ADDRESS}}", str(vendor.get("배송지", "")))
        .replace("{{VENDOR_PHONE}}", str(vendor.get("연락처", "")))
        .replace("{{ITEM_ROWS}}", rows_html)
        .replace("{{TOTAL_COUNT}}", base_app.fmt_int(total_count))
        .replace("{{TOTAL_QTY}}", base_app.fmt_int(total_qty))
        .replace("{{REQUEST_NOTE}}", request_note or "")
    )
    if not str(request_note or "").strip():
        html = re.sub(r'\s*<div class="request">\s*<b>요청사항</b><br>\s*-\s*</div>\s*', "\n", html, flags=re.DOTALL)
    return html


def create_excel_with_packaging(vendor, order_items, request_note, order_date=None):
    path = ORIGINAL_CREATE_EXCEL(vendor, order_items, request_note, order_date)
    wb = load_workbook(path)
    ws = wb.active
    start_row = 14
    ws.cell(start_row, 5, "포장단위")
    ws.cell(start_row, 6, "수량")
    for idx, item in enumerate(order_items, 1):
        row_num = start_row + idx
        ws.cell(row_num, 5, item.get("포장단위", item.get("단위", "")))
        ws.cell(row_num, 6, base_app.safe_int(item.get("수량", 0)))
    wb.save(path)
    return path


base_app.load_data = load_data_with_product_schema
base_app.save_products = save_products_with_schema
base_app.page_product_manage = page_product_manage_with_schema
base_app.render_order_html = render_order_html_with_packaging
base_app.create_excel = create_excel_with_packaging


if __name__ == "__main__":
    purchase.main()
