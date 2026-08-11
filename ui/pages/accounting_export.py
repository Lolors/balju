"""월마감 조회와 엑셀 내보내기."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _to_int(purchase_module, value) -> int:
    return purchase_module.to_int(value)


def _date_text(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    return "" if pd.isna(parsed) else parsed.strftime("%Y-%m-%d")


def _money(value) -> str:
    try:
        return f"{int(value):,}"
    except Exception:
        return "0"


def _build_frames(purchase_module, data, year: int, month: int):
    statements, statement_items, _, _ = purchase_module.load_purchase_data()
    orders = data["orders"].copy()

    if statements.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    working = statements.copy()
    working["_date"] = pd.to_datetime(working["명세서일자"], errors="coerce")
    working = working[
        (working["_date"].dt.year == int(year))
        & (working["_date"].dt.month == int(month))
    ].copy()
    if working.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    order_dates = {}
    if not orders.empty:
        for _, row in orders.iterrows():
            order_dates[str(row.get("발주ID", ""))] = _date_text(row.get("발주일시", ""))

    summary_rows = []
    detail_rows = []
    for _, statement in working.sort_values(["명세서일자", "등록일시"]).iterrows():
        statement_id = str(statement.get("명세서ID", ""))
        order_id = str(statement.get("발주ID", ""))
        vendor_name = str(statement.get("거래처명", ""))
        statement_number = str(statement.get("명세서번호", ""))
        statement_date = _date_text(statement.get("명세서일자", ""))
        order_date = order_dates.get(order_id, "")
        items = statement_items[
            statement_items["명세서ID"].astype(str) == statement_id
        ].copy()

        item_qty = 0
        product_amount = 0
        for _, item in items.iterrows():
            qty = _to_int(purchase_module, item.get("입고수량", 0))
            buy_price = _to_int(purchase_module, item.get("매입단가", 0))
            amount = _to_int(purchase_module, item.get("상품금액", qty * buy_price))
            item_qty += qty
            product_amount += amount
            detail_rows.append({
                "거래처": vendor_name,
                "발주일자": order_date,
                "발주번호": order_id,
                "거래명세서 번호": statement_number,
                "거래명세서 일자": statement_date,
                "정식제품명": str(item.get("정식제품명", "")),
                "규격": str(item.get("규격", "")),
                "수량": qty,
                "포장단위": str(item.get("포장단위", "") or item.get("단위", "")),
                "매입단가": buy_price,
                "금액": amount,
            })

        freight = _to_int(purchase_module, statement.get("운송비", 0))
        if freight:
            detail_rows.append({
                "거래처": vendor_name,
                "발주일자": order_date,
                "발주번호": order_id,
                "거래명세서 번호": statement_number,
                "거래명세서 일자": statement_date,
                "정식제품명": "배송비",
                "규격": "",
                "수량": "",
                "포장단위": "",
                "매입단가": "",
                "금액": freight,
            })

        summary_rows.append({
            "거래처": vendor_name,
            "발주일자": order_date,
            "발주번호": order_id,
            "거래명세서 번호": statement_number,
            "거래명세서 일자": statement_date,
            "품목 수": len(items),
            "입고수량 합계": item_qty,
            "상품금액": product_amount,
            "배송비": freight,
            "총 매입금액": product_amount + freight,
        })

    summary = pd.DataFrame(summary_rows)
    detail = pd.DataFrame(detail_rows)
    vendor = (
        summary.groupby("거래처", as_index=False)
        .agg(
            거래명세서수=("거래명세서 번호", "count"),
            상품금액=("상품금액", "sum"),
            배송비=("배송비", "sum"),
            총매입금액=("총 매입금액", "sum"),
        )
        if not summary.empty
        else pd.DataFrame()
    )
    return summary, detail, vendor


def _total_row(frame: pd.DataFrame, sum_columns: set[str]) -> dict:
    row = {column: "" for column in frame.columns}
    if len(frame.columns):
        row[frame.columns[0]] = "합계"
    for column in sum_columns:
        if column in frame.columns:
            row[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0).sum()
    return row


def _excel_bytes(month_key: str, summary: pd.DataFrame, detail: pd.DataFrame, vendor: pd.DataFrame) -> bytes:
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(
        name: str,
        frame: pd.DataFrame,
        money_columns: set[str],
        sum_columns: set[str],
        header_row: int = 1,
    ):
        ws = wb.create_sheet(name)
        headers = list(frame.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(header_row, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF7")
            cell.alignment = Alignment(horizontal="center")

        data_start = header_row + 1
        for row_idx, row in enumerate(frame.itertuples(index=False), data_start):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row_idx, col_idx, value)
                header = headers[col_idx - 1]
                if header in money_columns:
                    cell.number_format = "#,##0"
                cell.alignment = Alignment(
                    horizontal="left" if header in {"거래처", "정식제품명", "규격"} else "center"
                )

        total_row_index = data_start + len(frame)
        total_values = _total_row(frame, sum_columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(total_row_index, col_idx, total_values.get(header, ""))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F3F4F6")
            if header in money_columns:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center"
            )

        for idx, header in enumerate(headers, 1):
            values = [str(header)] + [str(v) for v in frame.iloc[:, idx - 1].tolist()[:200]]
            width = min(max(len(v) for v in values) + 3, 38)
            ws.column_dimensions[get_column_letter(idx)].width = max(width, 12)

        last_column = get_column_letter(max(1, len(headers)))
        data_end = max(header_row, data_start + len(frame) - 1)
        ws.auto_filter.ref = f"A{header_row}:{last_column}{data_end}"
        ws.freeze_panes = f"A{header_row + 1}"
        return ws

    summary_ws = add_sheet(
        "월마감 요약",
        summary,
        {"상품금액", "배송비", "총 매입금액"},
        {"입고수량 합계", "상품금액", "배송비", "총 매입금액"},
        header_row=3,
    )
    summary_ws["A1"] = f"{month_key}_메디풀_매입내역"
    summary_ws["A1"].font = Font(size=16, bold=True)
    summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(summary.columns)))
    summary_ws["A1"].alignment = Alignment(horizontal="center")

    add_sheet(
        "품목 상세",
        detail,
        {"매입단가", "금액"},
        {"수량", "금액"},
    )
    add_sheet(
        "거래처별 합계",
        vendor,
        {"상품금액", "배송비", "총매입금액"},
        {"거래명세서수", "상품금액", "배송비", "총매입금액"},
    )

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render(purchase_module, data) -> None:
    st = purchase_module.st
    st.markdown("## 월별 매입 현황")
    st.caption("거래명세서 일자 기준으로 월마감 자료를 조회하고 내려받습니다.")

    today = datetime.now()
    c1, c2 = st.columns(2)
    year = int(c1.number_input("연도", min_value=2020, max_value=2100, value=today.year, step=1))
    month = int(c2.number_input("월", min_value=1, max_value=12, value=today.month, step=1))
    month_key = f"{year:04d}-{month:02d}"

    summary, detail, vendor = _build_frames(purchase_module, data, year, month)
    if summary.empty:
        st.info("선택한 월의 거래명세서가 없습니다.")
        return

    product_amount = int(summary["상품금액"].sum())
    freight = int(summary["배송비"].sum())
    total = int(summary["총 매입금액"].sum())
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("거래명세서", f"{len(summary):,}건")
    m2.metric("상품금액", f"{product_amount:,}원")
    m3.metric("배송비", f"{freight:,}원")
    m4.metric("총 매입금액", f"{total:,}원")

    st.markdown("### 거래명세서별 월마감 요약")
    display = summary.copy()
    for col in ["상품금액", "배송비", "총 매입금액"]:
        display[col] = display[col].apply(_money)
    st.dataframe(display, use_container_width=True, hide_index=True)

    st.markdown("### 거래처별 합계")
    vendor_display = vendor.copy()
    for col in ["상품금액", "배송비", "총매입금액"]:
        vendor_display[col] = vendor_display[col].apply(_money)
    st.dataframe(vendor_display, use_container_width=True, hide_index=True)

    st.download_button(
        "월마감 엑셀 내려받기",
        data=_excel_bytes(month_key, summary, detail, vendor),
        file_name=f"{month_key}_메디풀_매입내역.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
