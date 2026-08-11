
import base64
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from rapidfuzz import fuzz

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


BASE = Path(__file__).parent
DATA = BASE / "data"
TEMPLATES = BASE / "templates"
ASSETS = BASE / "assets"
OUTPUT = BASE / "output"
EXCEL_OUTPUT = OUTPUT / "excel"
PDF_OUTPUT = OUTPUT / "pdf"

DATA.mkdir(exist_ok=True)
TEMPLATES.mkdir(exist_ok=True)
ASSETS.mkdir(exist_ok=True)
EXCEL_OUTPUT.mkdir(parents=True, exist_ok=True)
PDF_OUTPUT.mkdir(parents=True, exist_ok=True)

LOGO_FILE = ASSETS / "logo.png"
ROOT_LOGO_FILE = BASE / "logo.png"

COMPANY = {
    "상호": "주식회사 노투스팜",
    "대표": "노진국",
    "주소": "경기도 용인시 처인구 포곡읍 에버랜드로 57 4층",
    "연락처": "TEL. 02-6925-0406",
    "팩스": "FAX. 070-7507-2003",
    "등록번호": "728-88-01901",
}

VENDORS_FILE = DATA / "vendors.csv"
PRODUCTS_FILE = DATA / "products.csv"
ALIASES_FILE = DATA / "aliases.csv"
DRAFTS_FILE = DATA / "drafts.csv"
DRAFT_ITEMS_FILE = DATA / "draft_items.csv"
ORDERS_FILE = DATA / "orders.csv"
ORDER_ITEMS_FILE = DATA / "order_items.csv"
TEMPLATE_FILE = TEMPLATES / "purchase_order.html"


st.set_page_config(
    page_title="발주관리 시스템 v17",
    layout="wide",
    initial_sidebar_state="expanded",
)


def create_sample_files():
    if not VENDORS_FILE.exists():
        VENDORS_FILE.write_text(
            """거래처코드,거래처명,담당자,연락처,이메일,배송지
V001,대한약품,김대리,02-1234-5678,order@daehanpharm.co.kr,서울특별시 강남구 테헤란로 123 (1층 창고)
V002,메디풀,박과장,031-111-2222,order@medipool.co.kr,경기도 성남시 분당구 판교로 77
""",
            encoding="utf-8-sig",
        )

    if not PRODUCTS_FILE.exists():
        PRODUCTS_FILE.write_text(
            """제품코드,정식제품명,규격,단위
P001,엠마오리드카인크림,500g,EA
P010,메디알연고,100g,EA
P023,리도카인패치,50매/팩,EA
P030,소독용에탄올,500ml,EA
""",
            encoding="utf-8-sig",
        )

    if not ALIASES_FILE.exists():
        ALIASES_FILE.write_text(
            """거래처명,별칭,제품코드
전체,마취크림,P001
전체,엠마크림,P001
전체,리도카인크림,P001
전체,연고,P010
전체,패치,P023
전체,에탄올,P030
대한약품,마취,P001
""",
            encoding="utf-8-sig",
        )

    if not DRAFTS_FILE.exists():
        DRAFTS_FILE.write_text(
            "임시ID,작성일시,거래처명,요청사항,상태,총품목수,총수량\n",
            encoding="utf-8-sig",
        )

    if not DRAFT_ITEMS_FILE.exists():
        DRAFT_ITEMS_FILE.write_text(
            "임시ID,순번,제품코드,정식제품명,검색별칭,규격,단위,수량\n",
            encoding="utf-8-sig",
        )

    if not ORDERS_FILE.exists():
        ORDERS_FILE.write_text(
            "발주ID,발주일시,거래처명,요청사항,상태,총품목수,총수량\n",
            encoding="utf-8-sig",
        )

    if not ORDER_ITEMS_FILE.exists():
        ORDER_ITEMS_FILE.write_text(
            "발주ID,순번,제품코드,정식제품명,검색별칭,규격,단위,수량\n",
            encoding="utf-8-sig",
        )

    if not TEMPLATE_FILE.exists():
        TEMPLATE_FILE.write_text(DEFAULT_TEMPLATE, encoding="utf-8")


def read_csv(path, columns):
    if path.exists():
        df = pd.read_csv(path, dtype=str).fillna("")

        for col in columns:
            if col not in df.columns:
                df[col] = ""

        return df[columns]

    return pd.DataFrame(columns=columns)


def load_data():
    create_sample_files()

    vendors = read_csv(VENDORS_FILE, ["거래처코드", "거래처명", "담당자", "연락처", "이메일", "배송지"])
    products = read_csv(PRODUCTS_FILE, ["제품코드", "정식제품명", "규격", "단위"])
    aliases = read_csv(ALIASES_FILE, ["거래처명", "별칭", "제품코드"])
    drafts = read_csv(DRAFTS_FILE, ["임시ID", "작성일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    draft_items = read_csv(DRAFT_ITEMS_FILE, ["임시ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])
    orders = read_csv(ORDERS_FILE, ["발주ID", "발주일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    order_items = read_csv(ORDER_ITEMS_FILE, ["발주ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])

    return vendors, products, aliases, drafts, draft_items, orders, order_items


def fmt_int(value):
    try:
        return f"{int(float(value)):,}"
    except Exception:
        return "0"


def safe_int(value, default=0):
    try:
        return int(float(value))
    except Exception:
        return default


def get_logo_base64():
    logo = LOGO_FILE if LOGO_FILE.exists() else ROOT_LOGO_FILE
    if logo.exists():
        try:
            return base64.b64encode(logo.read_bytes()).decode("utf-8")
        except Exception:
            return ""
    return ""


def init_state():
    if "current_page" not in st.session_state:
        st.session_state.current_page = "발주 작성"

    if "order_items" not in st.session_state:
        st.session_state.order_items = []

    if "loaded_vendor_name" not in st.session_state:
        st.session_state.loaded_vendor_name = ""

    if "loaded_request_note" not in st.session_state:
        st.session_state.loaded_request_note = ""


def search_products(keyword, vendor_name, products, aliases):
    keyword = (keyword or "").strip()

    if not keyword:
        return pd.DataFrame()

    rows = []
    active = products.copy()
    alias_df = aliases[aliases["거래처명"].isin([vendor_name, "전체", ""])]

    for _, alias_row in alias_df.iterrows():
        alias = str(alias_row["별칭"])
        score = fuzz.partial_ratio(keyword, alias)

        if keyword in alias:
            score += 30

        if score >= 45:
            product = active[active["제품코드"] == alias_row["제품코드"]]

            if not product.empty:
                p = product.iloc[0]
                rows.append({
                    "별칭(검색어)": alias,
                    "제품코드": p["제품코드"],
                    "정식제품명": p["정식제품명"],
                    "규격": p["규격"],
                    "단위": p["단위"],
                    "점수": score + 10,
                })

    for _, p in active.iterrows():
        official = str(p["정식제품명"])
        score = fuzz.partial_ratio(keyword, official)

        if keyword in official:
            score += 30

        if score >= 45:
            rows.append({
                "별칭(검색어)": keyword,
                "제품코드": p["제품코드"],
                "정식제품명": p["정식제품명"],
                "규격": p["규격"],
                "단위": p["단위"],
                "점수": score,
            })

    if not rows:
        return pd.DataFrame()

    result = pd.DataFrame(rows)
    result = result.sort_values("점수", ascending=False)
    result = result.drop_duplicates("제품코드", keep="first")
    return result.reset_index(drop=True)


def normalize_order_items(rows):
    result = []

    for _, row in rows.iterrows():
        result.append({
            "제품코드": row.get("제품코드", ""),
            "정식제품명": row.get("정식제품명", ""),
            "검색별칭": row.get("검색별칭", ""),
            "규격": row.get("규격", ""),
            "단위": row.get("단위", ""),
            "수량": safe_int(row.get("수량", 0)),
        })

    return result


def calc_totals(order_items):
    total_count = len(order_items)
    total_qty = sum(safe_int(item.get("수량", 0)) for item in order_items)
    return total_count, total_qty


def get_latest_order_id(vendor_name, orders):
    if orders.empty:
        return ""

    vendor_orders = orders[orders["거래처명"] == vendor_name].copy()

    if vendor_orders.empty:
        return ""

    vendor_orders = vendor_orders.sort_values("발주일시", ascending=False)
    return vendor_orders.iloc[0]["발주ID"]


def get_frequent_products(vendor_name, orders, order_items_saved, limit=5):
    if orders.empty or order_items_saved.empty:
        return pd.DataFrame()

    vendor_order_ids = orders[orders["거래처명"] == vendor_name]["발주ID"].tolist()

    if not vendor_order_ids:
        return pd.DataFrame()

    rows = order_items_saved[order_items_saved["발주ID"].isin(vendor_order_ids)].copy()

    if rows.empty:
        return pd.DataFrame()

    rows["수량"] = pd.to_numeric(rows["수량"], errors="coerce").fillna(0).astype(int)

    grouped = (
        rows.groupby(["제품코드", "정식제품명", "검색별칭", "규격", "단위"], dropna=False)
        .agg(주문횟수=("발주ID", "count"), 최근수량=("수량", "last"), 누적수량=("수량", "sum"))
        .reset_index()
        .sort_values(["주문횟수", "누적수량"], ascending=False)
        .head(limit)
    )

    return grouped


def save_draft(vendor_name, request_note, order_items):
    drafts = read_csv(DRAFTS_FILE, ["임시ID", "작성일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    draft_items = read_csv(DRAFT_ITEMS_FILE, ["임시ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])

    draft_id = f"TEMP-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    total_count, total_qty = calc_totals(order_items)

    drafts = pd.concat([
        drafts,
        pd.DataFrame([{
            "임시ID": draft_id,
            "작성일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "거래처명": vendor_name,
            "요청사항": request_note,
            "상태": "임시저장",
            "총품목수": total_count,
            "총수량": total_qty,
        }])
    ], ignore_index=True)

    rows = []
    for idx, item in enumerate(order_items, 1):
        rows.append({
            "임시ID": draft_id,
            "순번": idx,
            "제품코드": item.get("제품코드", ""),
            "정식제품명": item.get("정식제품명", ""),
            "검색별칭": item.get("검색별칭", ""),
            "규격": item.get("규격", ""),
            "단위": item.get("단위", ""),
            "수량": safe_int(item.get("수량", 0)),
        })

    if rows:
        draft_items = pd.concat([draft_items, pd.DataFrame(rows)], ignore_index=True)

    drafts.to_csv(DRAFTS_FILE, index=False, encoding="utf-8-sig")
    draft_items.to_csv(DRAFT_ITEMS_FILE, index=False, encoding="utf-8-sig")

    return draft_id


def save_order(vendor_name, request_note, order_items):
    orders = read_csv(ORDERS_FILE, ["발주ID", "발주일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    order_items_df = read_csv(ORDER_ITEMS_FILE, ["발주ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])

    order_id = f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    total_count, total_qty = calc_totals(order_items)

    orders = pd.concat([
        orders,
        pd.DataFrame([{
            "발주ID": order_id,
            "발주일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "거래처명": vendor_name,
            "요청사항": request_note,
            "상태": "발주완료",
            "총품목수": total_count,
            "총수량": total_qty,
        }])
    ], ignore_index=True)

    rows = []
    for idx, item in enumerate(order_items, 1):
        rows.append({
            "발주ID": order_id,
            "순번": idx,
            "제품코드": item.get("제품코드", ""),
            "정식제품명": item.get("정식제품명", ""),
            "검색별칭": item.get("검색별칭", ""),
            "규격": item.get("규격", ""),
            "단위": item.get("단위", ""),
            "수량": safe_int(item.get("수량", 0)),
        })

    if rows:
        order_items_df = pd.concat([order_items_df, pd.DataFrame(rows)], ignore_index=True)

    orders.to_csv(ORDERS_FILE, index=False, encoding="utf-8-sig")
    order_items_df.to_csv(ORDER_ITEMS_FILE, index=False, encoding="utf-8-sig")

    return order_id


def delete_order(order_id):
    orders = read_csv(ORDERS_FILE, ["발주ID", "발주일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    order_items = read_csv(ORDER_ITEMS_FILE, ["발주ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])

    orders = orders[orders["발주ID"] != order_id]
    order_items = order_items[order_items["발주ID"] != order_id]

    orders.to_csv(ORDERS_FILE, index=False, encoding="utf-8-sig")
    order_items.to_csv(ORDER_ITEMS_FILE, index=False, encoding="utf-8-sig")


def delete_draft(draft_id):
    drafts = read_csv(DRAFTS_FILE, ["임시ID", "작성일시", "거래처명", "요청사항", "상태", "총품목수", "총수량"])
    draft_items = read_csv(DRAFT_ITEMS_FILE, ["임시ID", "순번", "제품코드", "정식제품명", "검색별칭", "규격", "단위", "수량"])

    drafts = drafts[drafts["임시ID"] != draft_id]
    draft_items = draft_items[draft_items["임시ID"] != draft_id]

    drafts.to_csv(DRAFTS_FILE, index=False, encoding="utf-8-sig")
    draft_items.to_csv(DRAFT_ITEMS_FILE, index=False, encoding="utf-8-sig")


def render_order_html(vendor, order_items, request_note, order_id=None, order_date=None):
    template = TEMPLATE_FILE.read_text(encoding="utf-8") if TEMPLATE_FILE.exists() else DEFAULT_TEMPLATE
    logo_b64 = get_logo_base64()

    order_id = order_id or f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    order_date = order_date or datetime.now().strftime("%Y-%m-%d")

    rows_html = ""
    for idx, item in enumerate(order_items, 1):
        rows_html += f"""
        <tr>
            <td>{idx}</td>
            <td>{item.get("제품코드", "")}</td>
            <td class="left">{item.get("정식제품명", "")}</td>
            <td>{item.get("규격", "")}</td>
            <td>{item.get("단위", "")}</td>
            <td>{fmt_int(item.get("수량", 0))}</td>
        </tr>
        """

    if not rows_html:
        rows_html = '<tr><td colspan="6" class="empty">발주 품목이 없습니다.</td></tr>'

    total_count, total_qty = calc_totals(order_items)

    if logo_b64:
        template = template.replace("{% if_logo %}", "").replace("{% else_logo %}", "<!--").replace("{% endif_logo %}", "-->")
    else:
        template = template.replace("{% if_logo %}", "<!--").replace("{% else_logo %}", "-->").replace("{% endif_logo %}", "")

    return (
        template
        .replace("{{LOGO_BASE64}}", logo_b64)
        .replace("{{ORDER_ID}}", order_id)
        .replace("{{ORDER_DATE}}", order_date)
        .replace("{{COMPANY_NAME}}", COMPANY["상호"])
        .replace("{{COMPANY_OWNER}}", COMPANY["대표"])
        .replace("{{COMPANY_ADDRESS}}", COMPANY["주소"])
        .replace("{{COMPANY_PHONE}}", COMPANY["연락처"])
        .replace("{{COMPANY_FAX}}", COMPANY["팩스"])
        .replace("{{COMPANY_REGNO}}", COMPANY["등록번호"])
        .replace("{{VENDOR_NAME}}", str(vendor.get("거래처명", "")))
        .replace("{{VENDOR_ADDRESS}}", str(vendor.get("배송지", "")))
        .replace("{{VENDOR_PHONE}}", str(vendor.get("연락처", "")))
        .replace("{{ITEM_ROWS}}", rows_html)
        .replace("{{TOTAL_COUNT}}", fmt_int(total_count))
        .replace("{{TOTAL_QTY}}", fmt_int(total_qty))
        .replace("{{REQUEST_NOTE}}", request_note or "")
    )


def create_excel(vendor, order_items, request_note, order_date=None):
    order_id = f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    path = EXCEL_OUTPUT / f"{order_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    for col, width in enumerate([8, 16, 32, 14, 10, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "발 주 서"
    ws["A1"].font = Font(size=24, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "발주번호"
    ws["B3"] = order_id
    ws["E3"] = "발주일자"
    ws["F3"] = order_date or datetime.now().strftime("%Y-%m-%d")

    ws.merge_cells("A5:C5")
    ws["A5"] = "발주자"
    ws["A5"].font = Font(bold=True)
    ws["A5"].fill = header_fill

    company_rows = [
        ("상호", COMPANY["상호"]),
        ("대표", COMPANY["대표"]),
        ("주소", COMPANY["주소"]),
        ("연락처", COMPANY["연락처"]),
        ("팩스", COMPANY["팩스"]),
        ("등록번호", COMPANY["등록번호"]),
    ]

    for r, (label, value) in enumerate(company_rows, 6):
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    ws.merge_cells("D5:F5")
    ws["D5"] = "납품처"
    ws["D5"].font = Font(bold=True)
    ws["D5"].fill = header_fill

    receiver_rows = [
        ("상호", vendor["거래처명"]),
        ("연락처", vendor["연락처"]),
        ("주소", vendor["배송지"]),
    ]

    for r, (label, value) in enumerate(receiver_rows, 6):
        ws[f"D{r}"] = label
        ws[f"E{r}"] = value
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

    start_row = 14
    headers = ["No.", "제품코드", "정식제품명", "규격", "단위", "수량"]

    for c, header in enumerate(headers, 1):
        cell = ws.cell(start_row, c, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(order_items, 1):
        row_num = start_row + idx
        values = [
            idx,
            item.get("제품코드", ""),
            item.get("정식제품명", ""),
            item.get("규격", ""),
            item.get("단위", ""),
            safe_int(item.get("수량", 0)),
        ]

        for c, value in enumerate(values, 1):
            cell = ws.cell(row_num, c, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 3 else "left")

    total_count, total_qty = calc_totals(order_items)
    total_row = start_row + len(order_items) + 2
    ws.cell(total_row, 4, "총 품목수")
    ws.cell(total_row, 5, total_count)
    ws.cell(total_row + 1, 4, "총 수량")
    ws.cell(total_row + 1, 5, total_qty)

    ws.cell(total_row + 3, 1, "요청사항")
    ws.cell(total_row + 4, 1, request_note or "")

    wb.save(path)
    return path


def get_pdf_font_name():
    candidates = [
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/malgunbd.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    ]

    for font_path in candidates:
        if Path(font_path).exists():
            try:
                pdfmetrics.registerFont(TTFont("KoreanFont", font_path))
                return "KoreanFont"
            except Exception:
                pass

    return "Helvetica"




def create_preview_image(vendor, order_items, request_note, order_date=None):
    """
    발주서 미리보기 HTML을 브라우저로 렌더링한 뒤,
    .paper 영역을 PNG 이미지로 캡쳐합니다.
    """
    image_path = PDF_OUTPUT / f"PO-PREVIEW-{datetime.now().strftime('%Y%m%d-%H%M%S')}.png"
    html = render_order_html(vendor, order_items, request_note, order_date=order_date)

    from playwright.sync_api import sync_playwright

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir = Path(temp_dir)
        html_path = temp_dir / "purchase_order_preview.html"

        html_path.write_text(html, encoding="utf-8")

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(
                viewport={"width": 1200, "height": 1600},
                device_scale_factor=2,
            )
            page.goto(html_path.as_uri(), wait_until="networkidle")
            page.locator(".paper").screenshot(path=str(image_path))
            browser.close()

    return image_path


def create_preview_pdf(vendor, order_items, request_note, order_date=None):
    """
    미리보기 캡쳐 PNG를 A4 PDF 안에 그대로 넣습니다.
    기존 ReportLab 표 생성 방식으로 fallback하지 않습니다.
    """
    pdf_path = PDF_OUTPUT / f"PO-PREVIEW-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    png_path = create_preview_image(vendor, order_items, request_note, order_date=order_date)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    page_w, page_h = A4
    margin = 18

    img = ImageReader(str(png_path))
    img_w, img_h = img.getSize()

    max_w = page_w - margin * 2
    max_h = page_h - margin * 2
    scale = min(max_w / img_w, max_h / img_h)

    draw_w = img_w * scale
    draw_h = img_h * scale
    x = (page_w - draw_w) / 2
    y = page_h - margin - draw_h

    c.drawImage(str(png_path), x, y, width=draw_w, height=draw_h)
    c.save()

    return pdf_path



def create_pdf(vendor, order_items, request_note, order_date=None):
    order_id = f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    path = PDF_OUTPUT / f"{order_id}.pdf"
    font_name = get_pdf_font_name()

    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    x_margin = 40
    y = height - 45

    logo = LOGO_FILE if LOGO_FILE.exists() else ROOT_LOGO_FILE
    if logo.exists():
        try:
            c.drawImage(str(logo), x_margin, y - 18, width=145, height=40, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    c.setFont(font_name, 24)
    c.drawCentredString(width / 2, y, "발 주 서")

    y -= 34
    c.setFont(font_name, 9)
    c.drawString(x_margin, y, f"발주번호: {order_id}")
    c.drawRightString(width - x_margin, y, f"발주일자: {order_date or datetime.now().strftime('%Y-%m-%d')}")

    y -= 25
    c.line(x_margin, y, width - x_margin, y)
    y -= 20

    c.setFont(font_name, 12)
    c.drawString(x_margin, y, "발주자")
    c.drawString(width / 2 + 10, y, "납품처")

    y -= 18
    c.setFont(font_name, 9)

    left_rows = [
        ("상호", COMPANY["상호"]),
        ("대표", COMPANY["대표"]),
        ("주소", COMPANY["주소"]),
        ("연락처", COMPANY["연락처"]),
        ("팩스", COMPANY["팩스"]),
        ("등록번호", COMPANY["등록번호"]),
    ]

    right_rows = [
        ("상호", vendor["거래처명"]),
        ("연락처", vendor["연락처"]),
        ("주소", vendor["배송지"]),
    ]

    y_start = y
    for label, value in left_rows:
        c.drawString(x_margin, y, f"{label}: {value}")
        y -= 15

    y = y_start
    for label, value in right_rows:
        c.drawString(width / 2 + 10, y, f"{label}: {value}")
        y -= 15

    y -= 25
    c.line(x_margin, y, width - x_margin, y)
    y -= 22

    headers = ["No.", "제품코드", "제품명", "규격", "단위", "수량"]
    col_x = [40, 75, 145, 320, 390, 455]

    c.setFont(font_name, 9)
    for i, header in enumerate(headers):
        c.drawString(col_x[i], y, header)

    y -= 10
    c.line(x_margin, y, width - x_margin, y)
    y -= 17

    for idx, item in enumerate(order_items, 1):
        values = [
            idx,
            item.get("제품코드", ""),
            item.get("정식제품명", ""),
            item.get("규격", ""),
            item.get("단위", ""),
            fmt_int(item.get("수량", 0)),
        ]

        for i, value in enumerate(values):
            c.drawString(col_x[i], y, str(value)[:28])

        y -= 18

        if y < 100:
            c.showPage()
            c.setFont(font_name, 9)
            y = height - 50

    total_count, total_qty = calc_totals(order_items)
    y -= 5
    c.line(x_margin, y, width - x_margin, y)
    y -= 22

    c.setFont(font_name, 11)
    c.drawRightString(width - x_margin, y, f"총 품목수 : {fmt_int(total_count)}건    총 수량 : {fmt_int(total_qty)} EA")

    y -= 35
    c.setFont(font_name, 10)
    c.drawString(x_margin, y, "요청사항")
    y -= 18
    c.drawString(x_margin, y, f"- {request_note or ''}")

    c.save()
    return path


def inject_css():
    st.markdown("""
<style>
.block-container {
    padding-top: 2.6rem;
    padding-left: 1.2rem;
    padding-right: 1.2rem;
    max-width: 1800px;
}

[data-testid="stSidebar"] {
    background: #ffffff;
    border-right: 1px solid #e5e7eb;
}

.sidebar-title {
    font-size: 22px;
    font-weight: 900;
    margin-bottom: 22px;
    padding-top: 8px;
}

.sidebar-section {
    margin-top: 24px;
    margin-bottom: 8px;
    font-weight: 900;
}

.help-box {
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 14px;
    margin-top: 40px;
    font-size: 13px;
}

section[data-testid="stSidebar"] .stButton button {
    justify-content: flex-start;
    text-align: left;
    padding-left: 18px;
    height: 52px;
    border-radius: 8px;
    font-size: 15px;
    font-weight: 800;
    border: 0;
    box-shadow: none;
    white-space: nowrap !important;
}

section[data-testid="stSidebar"] .stButton button[kind="primary"] {
    background: #0b5cff;
    color: white;
    border-color: #0b5cff;
}

section[data-testid="stSidebar"] .stButton button[kind="secondary"] {
    background: transparent;
    color: #111827;
}

section[data-testid="stSidebar"] .stButton button[kind="secondary"]:hover {
    background: #f3f6fb;
    color: #0b5cff;
}

section[data-testid="stSidebar"] hr {
    margin: 22px 0;
    border: none;
    border-top: 1px solid #e5e7eb;
}

.page-head {
    display: flex;
    align-items: center;
    gap: 24px;
    margin-top: 0;
    margin-bottom: 22px;
    min-height: 54px;
    padding-top: 8px;
    overflow: visible;
    flex-wrap: wrap;
}

.page-head-title {
    font-size: 23px;
    font-weight: 900;
    line-height: 1.6;
    padding-top: 6px;
    padding-bottom: 4px;
    overflow: visible;
}

.page-head-desc {
    color: #6b7280;
    font-size: 14px;
    line-height: 1.6;
    padding-top: 8px;
    overflow: visible;
}

.card-title {
    color: #0052cc;
    font-size: 18px;
    font-weight: 900;
    padding-bottom: 12px;
    border-bottom: 1px solid #e5e7eb;
    margin-bottom: 13px;
}

.vendor-card {
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 18px;
    min-height: 215px;
}

.vendor-name {
    font-size: 20px;
    font-weight: 900;
    margin-bottom: 18px;
}

.info-row {
    display: grid;
    grid-template-columns: 72px 1fr;
    gap: 10px;
    margin: 10px 0;
    font-size: 14px;
}

.info-label {
    color: #4b5563;
    font-weight: 800;
}

.hint {
    background: #eaf2ff;
    color: #0b5cff;
    padding: 9px 10px;
    border-radius: 7px;
    font-size: 13px;
    margin: 9px 0 14px 0;
}

.notice {
    background: #eaf2ff;
    color: #0b5cff;
    padding: 9px 10px;
    border-radius: 7px;
    font-size: 13px;
    margin-top: 8px;
}

.preview-title {
    font-size: 18px;
    font-weight: 900;
}

button {
    white-space: nowrap !important;
}

.preview-title + div {
    margin-top: 0;
}

button[title="엑셀 내려받기"],
button[title="PDF 생성"],
button[title="PDF 내려받기"],
button[title="PNG 생성"],
button[title="PNG 내려받기"] {
    min-width: 54px !important;
}

</style>
""", unsafe_allow_html=True)


def render_sidebar():
    def menu_button(name):
        is_active = st.session_state.current_page == name
        button_type = "primary" if is_active else "secondary"

        if st.sidebar.button(name, key=f"menu_{name}", use_container_width=True, type=button_type):
            st.session_state.current_page = name
            st.rerun()

    st.sidebar.markdown('<div class="sidebar-title">발주관리 시스템</div>', unsafe_allow_html=True)

    menu_button("발주 작성")
    menu_button("임시저장 목록")
    menu_button("발주서 목록")
    menu_button("최근 발주 내역")

    st.sidebar.markdown("<hr>", unsafe_allow_html=True)
    st.sidebar.markdown('<div class="sidebar-section">기초 관리</div>', unsafe_allow_html=True)

    menu_button("거래처 관리")
    menu_button("제품 관리")
    menu_button("별칭 관리")

    st.sidebar.markdown("<hr>", unsafe_allow_html=True)
    st.sidebar.markdown('<div class="sidebar-section">통계</div>', unsafe_allow_html=True)

    menu_button("발주 통계")
    menu_button("품목별 발주 통계")


    return st.session_state.current_page



def render_purchase_preview(vendor, order_items, request_note, order_date=None):
    top_left, top_excel, top_pdf, top_png = st.columns([2.6, 0.55, 0.55, 0.55])

    with top_left:
        st.markdown('<div class="preview-title">발주서 미리보기</div>', unsafe_allow_html=True)

    # 엑셀은 가벼워서 즉시 다운로드 버튼으로 제공
    with top_excel:
        excel_path = create_excel(vendor, order_items, request_note, order_date)
        with open(excel_path, "rb") as f:
            st.download_button(
                "XLS",
                f,
                file_name=excel_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"preview_xls_{len(order_items)}_{order_date}",
                help="엑셀 내려받기",
            )

    # PDF/PNG는 캡쳐 생성이 필요해서 첫 클릭은 생성, 생성 후 같은 아이콘이 다운로드 버튼으로 바뀜
    with top_pdf:
        pdf_saved_path = st.session_state.get("preview_pdf_path")
        if pdf_saved_path and Path(pdf_saved_path).exists():
            with open(pdf_saved_path, "rb") as f:
                st.download_button(
                    "PDF",
                    f,
                    file_name=Path(pdf_saved_path).name,
                    mime="application/pdf",
                    use_container_width=True,
                    key="preview_pdf_download_ready",
                    help="PDF 내려받기",
                )
        else:
            if st.button("PDF", use_container_width=True, key="preview_pdf_generate", help="PDF 생성"):
                try:
                    with st.spinner("PDF 생성 중..."):
                        pdf_path = create_preview_pdf(vendor, order_items, request_note, order_date)
                    st.session_state["preview_pdf_path"] = str(pdf_path)
                    st.session_state.pop("preview_capture_error", None)
                    st.rerun()
                except Exception as e:
                    st.session_state["preview_capture_error"] = str(e)

    with top_png:
        png_saved_path = st.session_state.get("preview_png_path")
        if png_saved_path and Path(png_saved_path).exists():
            with open(png_saved_path, "rb") as f:
                st.download_button(
                    "PNG",
                    f,
                    file_name=Path(png_saved_path).name,
                    mime="image/png",
                    use_container_width=True,
                    key="preview_png_download_ready",
                    help="PNG 내려받기",
                )
        else:
            if st.button("PNG", use_container_width=True, key="preview_png_generate", help="PNG 생성"):
                try:
                    with st.spinner("PNG 생성 중..."):
                        png_path = create_preview_image(vendor, order_items, request_note, order_date)
                    st.session_state["preview_png_path"] = str(png_path)
                    st.session_state.pop("preview_capture_error", None)
                    st.rerun()
                except Exception as e:
                    st.session_state["preview_capture_error"] = str(e)

    capture_error = st.session_state.get("preview_capture_error")
    if capture_error:
        st.error(
            "PDF/PNG 캡쳐 생성에 실패했습니다. 대부분은 앱을 실행한 Python과 명령어를 실행한 Python이 다른 경우입니다.\n\n"
            "앱 실행에 쓰는 같은 CMD에서 아래 명령을 실행하세요.\n"
            "python -m pip install playwright\n"
            "python -m playwright install chromium\n\n"
            f"오류내용: {capture_error}"
        )

    html = render_order_html(vendor, order_items, request_note, order_date=order_date)
    components.html(html, height=790, scrolling=True)



def page_drafts(drafts, draft_items):
    st.markdown("## 임시저장 목록")

    if drafts.empty:
        st.info("임시저장된 발주서가 없습니다.")
        return

    view = drafts.copy()
    view["총품목수"] = view["총품목수"].apply(fmt_int)
    view["총수량"] = view["총수량"].apply(fmt_int)

    st.dataframe(view, use_container_width=True, hide_index=True)

    selected_draft = st.selectbox("불러올 임시저장", view["임시ID"].tolist())

    c1, c2 = st.columns(2)

    with c1:
        if st.button("선택한 임시저장 불러오기", use_container_width=True):
            selected_header = drafts[drafts["임시ID"] == selected_draft].iloc[0]
            rows = draft_items[draft_items["임시ID"] == selected_draft].copy()

            if not rows.empty:
                st.session_state.order_items = normalize_order_items(rows)
                st.session_state.loaded_vendor_name = selected_header["거래처명"]
                st.session_state.loaded_request_note = selected_header.get("요청사항", "")
                st.session_state.current_page = "발주 작성"
                st.rerun()

    with c2:
        if st.button("선택한 임시저장 삭제", use_container_width=True):
            delete_draft(selected_draft)
            st.success("삭제했습니다.")
            st.rerun()


def page_orders(vendors, orders, order_items_saved):
    st.markdown("## 발주서 목록")

    if orders.empty:
        st.info("발주 완료 처리된 발주서가 없습니다.")
        return

    view = orders.copy()
    view["총품목수"] = view["총품목수"].apply(fmt_int)
    view["총수량"] = view["총수량"].apply(fmt_int)

    st.dataframe(view, use_container_width=True, hide_index=True)

    selected_order = st.selectbox("상세 확인할 발주서", orders["발주ID"].tolist())
    selected_header = orders[orders["발주ID"] == selected_order].iloc[0]
    detail = order_items_saved[order_items_saved["발주ID"] == selected_order].copy()

    if not detail.empty:
        detail_view = detail[["제품코드", "정식제품명", "규격", "단위", "수량"]].copy()
        detail_view["수량"] = detail_view["수량"].apply(fmt_int)

        st.markdown("### 발주 품목")
        st.dataframe(detail_view, use_container_width=True, hide_index=True)

        vendor_row = vendors[vendors["거래처명"] == selected_header["거래처명"]]

        c1, c2, c3, c4, c5 = st.columns(5)

        with c1:
            if st.button("수정", use_container_width=True):
                st.session_state.order_items = normalize_order_items(detail)
                st.session_state.loaded_vendor_name = selected_header["거래처명"]
                st.session_state.loaded_request_note = selected_header.get("요청사항", "")
                st.session_state.current_page = "발주 작성"
                st.rerun()

        with c2:
            if st.button("복사", use_container_width=True):
                st.session_state.order_items = normalize_order_items(detail)
                st.session_state.loaded_vendor_name = selected_header["거래처명"]
                st.session_state.loaded_request_note = selected_header.get("요청사항", "")
                st.session_state.current_page = "발주 작성"
                st.rerun()

        if not vendor_row.empty:
            vendor_for_export = vendor_row.iloc[0]
            items_for_export = normalize_order_items(detail)
            note_for_export = selected_header.get("요청사항", "")

            with c3:
                if st.button("엑셀 생성", use_container_width=True):
                    path = create_excel(vendor_for_export, items_for_export, note_for_export)
                    st.session_state[f"excel_path_{selected_order}"] = str(path)

                excel_path = st.session_state.get(f"excel_path_{selected_order}")
                if excel_path and Path(excel_path).exists():
                    with open(excel_path, "rb") as f:
                        st.download_button("엑셀 다운로드", f, file_name=Path(excel_path).name, use_container_width=True)

            with c4:
                if st.button("PDF 생성", use_container_width=True):
                    path = create_pdf(vendor_for_export, items_for_export, note_for_export)
                    st.session_state[f"pdf_path_{selected_order}"] = str(path)

                pdf_path = st.session_state.get(f"pdf_path_{selected_order}")
                if pdf_path and Path(pdf_path).exists():
                    with open(pdf_path, "rb") as f:
                        st.download_button("PDF 다운로드", f, file_name=Path(pdf_path).name, mime="application/pdf", use_container_width=True)

        with c5:
            if st.button("삭제", use_container_width=True):
                delete_order(selected_order)
                st.success("삭제했습니다.")
                st.rerun()


def page_recent_orders(orders):
    st.markdown("## 최근 발주 내역")

    if orders.empty:
        st.info("발주 이력이 없습니다.")
        return

    view = orders.copy().sort_values("발주일시", ascending=False).head(30)
    view["총품목수"] = view["총품목수"].apply(fmt_int)
    view["총수량"] = view["총수량"].apply(fmt_int)
    st.dataframe(view, use_container_width=True, hide_index=True)



def save_vendors(df):
    columns = ["거래처코드", "거래처명", "담당자", "연락처", "이메일", "배송지"]
    df = df[columns].fillna("")
    df.to_csv(VENDORS_FILE, index=False, encoding="utf-8-sig")


def save_products(df):
    columns = ["제품코드", "정식제품명", "규격", "단위"]

    for col in columns:
        if col not in df.columns:
            df[col] = ""

    df = df[columns].fillna("")
    df.to_csv(PRODUCTS_FILE, index=False, encoding="utf-8-sig")


def save_aliases(df):
    columns = ["거래처명", "별칭", "제품코드"]
    df = df[columns].fillna("")
    df.to_csv(ALIASES_FILE, index=False, encoding="utf-8-sig")


def make_next_code(existing_codes, prefix):
    numbers = []

    for code in existing_codes:
        text = str(code).replace(prefix, "")
        if text.isdigit():
            numbers.append(int(text))

    next_no = max(numbers) + 1 if numbers else 1
    return f"{prefix}{next_no:03d}"


def page_vendor_manage(vendors):
    st.markdown("## 거래처 관리")
    st.caption("거래처명, 담당자, 연락처, 납품처 주소를 관리합니다.")

    with st.container(border=True):
        st.markdown("### 신규 거래처 추가")

        default_code = make_next_code(vendors["거래처코드"].tolist(), "V")

        c1, c2 = st.columns(2)
        with c1:
            new_name = st.text_input("거래처명")
            new_address = st.text_input("납품처 주소")
        with c2:
            new_manager = st.text_input("담당자")
            new_phone = st.text_input("연락처")

        if st.button("거래처 추가", type="primary", use_container_width=True):
            if not new_name.strip():
                st.warning("거래처명을 입력하세요.")
            else:
                new_row = pd.DataFrame([{
                    "거래처코드": default_code,
                    "거래처명": new_name.strip(),
                    "담당자": new_manager.strip(),
                    "연락처": new_phone.strip(),
                    "이메일": "",
                    "배송지": new_address.strip(),
                }])
                vendors = pd.concat([vendors, new_row], ignore_index=True)
                save_vendors(vendors)
                st.success("거래처를 추가했습니다.")
                st.rerun()

    st.markdown("### 거래처 목록 수정")

    if vendors.empty:
        st.info("등록된 거래처가 없습니다.")
        return

    edit_df = vendors.copy()
    edit_df["삭제"] = False
    edit_df = edit_df[["삭제", "거래처명", "담당자", "연락처", "배송지"]]

    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "삭제": st.column_config.CheckboxColumn("삭제"),
            "배송지": st.column_config.TextColumn("납품처 주소"),
        },
        key="vendor_editor",
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("거래처 수정 저장", use_container_width=True):
            clean = edited[edited["삭제"] != True].drop(columns=["삭제"]).copy()

            # 화면에는 숨긴 거래처코드/이메일을 자동 보정
            clean["거래처코드"] = [
                make_next_code([], "V") if idx is None else ""
                for idx in range(len(clean))
            ]

            old_codes = vendors["거래처코드"].tolist()
            restored_codes = []

            for i, _ in clean.iterrows():
                if i in vendors.index:
                    restored_codes.append(vendors.loc[i, "거래처코드"])
                else:
                    restored_codes.append(make_next_code(old_codes + restored_codes, "V"))

            clean["거래처코드"] = restored_codes
            clean["이메일"] = ""

            clean = clean[["거래처코드", "거래처명", "담당자", "연락처", "이메일", "배송지"]]
            save_vendors(clean)
            st.success("거래처 정보를 저장했습니다.")
            st.rerun()

    with c2:
        if st.button("선택 거래처 삭제", use_container_width=True):
            clean = edited[edited["삭제"] != True].drop(columns=["삭제"]).copy()

            old_codes = vendors["거래처코드"].tolist()
            restored_codes = []

            for i, _ in clean.iterrows():
                if i in vendors.index:
                    restored_codes.append(vendors.loc[i, "거래처코드"])
                else:
                    restored_codes.append(make_next_code(old_codes + restored_codes, "V"))

            clean["거래처코드"] = restored_codes
            clean["이메일"] = ""
            clean = clean[["거래처코드", "거래처명", "담당자", "연락처", "이메일", "배송지"]]
            save_vendors(clean)
            st.success("선택한 거래처를 삭제했습니다.")
            st.rerun()


def page_product_manage(products):
    st.markdown("## 제품 관리")
    st.caption("제품코드, 정식제품명, 규격, 단위를 관리합니다. 단가/금액은 사용하지 않습니다.")

    with st.container(border=True):
        st.markdown("### 신규 제품 추가")

        c1, c2, c3 = st.columns(3)
        with c1:
            new_code = st.text_input("제품코드", value="", placeholder="업체 고유코드가 있으면 입력")
            new_name = st.text_input("정식제품명")
        with c2:
            new_spec = st.text_input("규격")
            new_unit = st.text_input("단위", value="EA")
        if st.button("제품 추가", type="primary", use_container_width=True):
            if not new_name.strip():
                st.warning("정식제품명을 입력하세요.")
            elif new_code.strip() and new_code.strip() in products["제품코드"].tolist():
                st.warning("이미 존재하는 제품코드입니다.")
            else:
                new_row = pd.DataFrame([{
                    "제품코드": new_code.strip(),
                    "정식제품명": new_name.strip(),
                    "규격": new_spec.strip(),
                    "단위": new_unit.strip(),
                }])
                products = pd.concat([products, new_row], ignore_index=True)
                save_products(products)
                st.success("제품을 추가했습니다.")
                st.rerun()

    st.markdown("### 제품 목록 수정")

    if products.empty:
        st.info("등록된 제품이 없습니다.")
        return

    edit_df = products.copy()
    edit_df["삭제"] = False
    edit_df = edit_df[["삭제", "제품코드", "정식제품명", "규격", "단위"]]

    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "삭제": st.column_config.CheckboxColumn("삭제"),
        },
        key="product_editor",
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("제품 수정 저장", use_container_width=True):
            clean = edited[edited["삭제"] != True].drop(columns=["삭제"])
            save_products(clean)
            st.success("제품 정보를 저장했습니다.")
            st.rerun()

    with c2:
        if st.button("선택 제품 삭제", use_container_width=True):
            clean = edited[edited["삭제"] != True].drop(columns=["삭제"])
            save_products(clean)
            st.success("선택한 제품을 삭제했습니다.")
            st.rerun()


def page_alias_manage(vendors, products, aliases):
    st.markdown("## 별칭 관리")
    st.caption("거래처가 말하는 약칭/별칭을 정식 제품에 연결합니다.")

    vendor_options = ["전체"] + vendors["거래처명"].drop_duplicates().tolist()

    with st.container(border=True):
        st.markdown("### 신규 별칭 추가")

        c1, c2, c3 = st.columns([1, 1, 2])

        with c1:
            new_vendor = st.selectbox("거래처명", vendor_options)

        with c2:
            new_alias = st.text_input("별칭")

        with c3:
            product_keyword = st.text_input("연결제품 검색", placeholder="제품명/제품코드 일부 입력")

            product_candidates = products.copy()

            if product_keyword.strip():
                mask = (
                    product_candidates["제품코드"].astype(str).str.contains(product_keyword, case=False, na=False)
                    | product_candidates["정식제품명"].astype(str).str.contains(product_keyword, case=False, na=False)
                    | product_candidates["규격"].astype(str).str.contains(product_keyword, case=False, na=False)
                )
                product_candidates = product_candidates[mask]

            product_candidates = product_candidates.head(30).copy()

            if product_candidates.empty:
                selected_product_label = ""
                st.caption("검색 결과가 없습니다.")
            else:
                product_options = [
                    f'{row["제품코드"]} | {row["정식제품명"]} ({row["규격"]} / {row["단위"]})'
                    for _, row in product_candidates.iterrows()
                ]

                selected_product_label = st.selectbox("연결제품", product_options)

        if st.button("별칭 추가", type="primary", use_container_width=True):
            if not new_alias.strip():
                st.warning("별칭을 입력하세요.")
            elif not selected_product_label:
                st.warning("연결할 제품을 선택하세요.")
            else:
                product_code = selected_product_label.split("|")[0].strip()
                product_name = selected_product_label.split("|")[1].split("(")[0].strip() if "|" in selected_product_label else ""

                # 제품코드가 공백인 제품도 있을 수 있으므로 제품명으로 재확인
                if not product_code:
                    selected_rows = products[products["정식제품명"] == product_name]
                    product_code = selected_rows.iloc[0]["제품코드"] if not selected_rows.empty else ""

                duplicate = aliases[
                    (aliases["거래처명"] == new_vendor)
                    & (aliases["별칭"] == new_alias.strip())
                    & (aliases["제품코드"] == product_code)
                ]

                if not duplicate.empty:
                    st.warning("이미 등록된 별칭입니다.")
                else:
                    new_row = pd.DataFrame([{
                        "거래처명": new_vendor,
                        "별칭": new_alias.strip(),
                        "제품코드": product_code,
                    }])
                    aliases = pd.concat([aliases, new_row], ignore_index=True)
                    save_aliases(aliases)
                    st.success("별칭을 추가했습니다.")
                    st.rerun()

    st.markdown("### 별칭 목록 수정")

    if aliases.empty:
        st.info("등록된 별칭이 없습니다.")
        return

    merged = aliases.merge(
        products[["제품코드", "정식제품명", "규격", "단위"]],
        on="제품코드",
        how="left",
    )

    view = merged[["거래처명", "별칭", "제품코드", "정식제품명", "규격", "단위"]].copy()
    view["삭제"] = False
    view = view[["삭제", "거래처명", "별칭", "제품코드", "정식제품명", "규격", "단위"]]

    edited = st.data_editor(
        view,
        use_container_width=True,
        hide_index=True,
        disabled=["정식제품명", "규격", "단위"],
        column_config={
            "삭제": st.column_config.CheckboxColumn("삭제"),
            "거래처명": st.column_config.SelectboxColumn("거래처명", options=vendor_options),
            "제품코드": st.column_config.TextColumn("제품코드"),
        },
        key="alias_editor",
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("별칭 수정 저장", use_container_width=True):
            clean = edited[edited["삭제"] != True][["거래처명", "별칭", "제품코드"]]
            save_aliases(clean)
            st.success("별칭 정보를 저장했습니다.")
            st.rerun()

    with c2:
        if st.button("선택 별칭 삭제", use_container_width=True):
            clean = edited[edited["삭제"] != True][["거래처명", "별칭", "제품코드"]]
            save_aliases(clean)
            st.success("선택한 별칭을 삭제했습니다.")
            st.rerun()



def page_placeholder(title):
    st.markdown(f"## {title}")
    st.info("이 메뉴는 다음 단계에서 연결하면 됩니다.")


def page_write(vendors, products, aliases, drafts, draft_items, orders, order_items_saved):
    st.markdown(
        """
<div class="page-head">
    <div class="page-head-title">발주 작성</div>
    <div class="page-head-desc">거래처를 선택하고 제품을 검색하여 발주 목록을 작성하세요.</div>
</div>
""",
        unsafe_allow_html=True,
    )

    left, right = st.columns([1.12, 1], gap="medium")

    with left:
        vendor_col, search_col = st.columns([0.85, 1.15], gap="small")

        with vendor_col:
            with st.container(border=True):
                st.markdown('<div class="card-title">1. 거래처 선택</div>', unsafe_allow_html=True)

                vendor_names = vendors["거래처명"].tolist()
                default_vendor = st.session_state.get("loaded_vendor_name") or vendor_names[0]
                default_index = vendor_names.index(default_vendor) if default_vendor in vendor_names else 0

                vendor_name = st.selectbox("거래처", vendor_names, index=default_index, label_visibility="collapsed")
                vendor = vendors[vendors["거래처명"] == vendor_name].iloc[0]

                st.markdown(
                    f"""
<div class="vendor-card">
    <div class="vendor-name">{vendor["거래처명"]}</div>
    <div class="info-row"><div class="info-label">담당자</div><div>{vendor["담당자"]}</div></div>
    <div class="info-row"><div class="info-label">연락처</div><div>{vendor["연락처"]}</div></div>
    <div class="info-row"><div class="info-label">이메일</div><div>{vendor["이메일"]}</div></div>
    <div class="info-row"><div class="info-label">주소</div><div>{vendor["배송지"]}</div></div>
</div>
""",
                    unsafe_allow_html=True,
                )

                latest_order_id = get_latest_order_id(vendor["거래처명"], orders)
                if latest_order_id:
                    if st.button("최근 발주 복사", use_container_width=True):
                        recent_rows = order_items_saved[order_items_saved["발주ID"] == latest_order_id].copy()
                        if not recent_rows.empty:
                            st.session_state.order_items = normalize_order_items(recent_rows)
                            st.success(f"최근 발주를 불러왔습니다: {latest_order_id}")
                            st.rerun()
                else:
                    st.caption("최근 발주 이력이 없습니다.")

        with search_col:
            with st.container(border=True):
                st.markdown(
                    '<div class="card-title">2. 제품 검색 <span style="font-size:14px;">(별칭/약칭으로 검색)</span></div>',
                    unsafe_allow_html=True,
                )

                keyword = st.text_input("검색", value="마취크림", placeholder="예: 마취크림", label_visibility="collapsed")

                st.markdown(
                    '<div class="hint">별칭(약칭)으로 검색해도 정식 제품명이 발주서에 적용됩니다.</div>',
                    unsafe_allow_html=True,
                )

                result = search_products(keyword, vendor["거래처명"], products, aliases)

                selected = None

                if result.empty:
                    st.warning("검색 결과가 없습니다.")
                else:
                    st.markdown(
                        '<div style="font-weight:900;color:#0052cc;margin:8px 0;">검색 결과</div>',
                        unsafe_allow_html=True,
                    )

                    show = result[["별칭(검색어)", "정식제품명", "제품코드", "규격"]].copy()

                    st.dataframe(show, use_container_width=True, hide_index=True, height=145)

                    selected_index = st.selectbox(
                        "선택",
                        list(result.index),
                        format_func=lambda i: f'{result.loc[i, "별칭(검색어)"]} → {result.loc[i, "정식제품명"]}',
                        label_visibility="collapsed",
                    )

                    selected = result.loc[selected_index]

                frequent_products = get_frequent_products(vendor["거래처명"], orders, order_items_saved, limit=5)

                if not frequent_products.empty:
                    st.markdown(
                        '<div style="font-weight:900;color:#0052cc;margin:12px 0 8px 0;">자주 주문하는 품목</div>',
                        unsafe_allow_html=True,
                    )

                    for idx, row in frequent_products.iterrows():
                        label = f'{row["정식제품명"]} / 최근수량 {fmt_int(row["최근수량"])} {row["단위"]}'

                        if st.button(label, use_container_width=True, key=f"freq_{vendor['거래처명']}_{row['제품코드']}_{idx}"):
                            qty = safe_int(row["최근수량"], 1) or 1

                            st.session_state.order_items.append({
                                "제품코드": row["제품코드"],
                                "정식제품명": row["정식제품명"],
                                "검색별칭": row["검색별칭"],
                                "규격": row["규격"],
                                "단위": row["단위"],
                                "수량": qty,
                            })

                            st.rerun()

        with st.container(border=True):
            if selected is not None:
                st.caption("선택된 제품")
                product_info_col, qty_col = st.columns([3, 2], gap="small")

                with product_info_col:
                    st.markdown(
                        f'**{selected["정식제품명"]} ({selected["제품코드"]})** &nbsp;&nbsp; `{selected["규격"]}`',
                        unsafe_allow_html=True,
                    )

                with qty_col:
                    qty = st.number_input("수량", min_value=1, value=10, step=1)

                    if st.button("추가", type="primary", use_container_width=True):
                        st.session_state.order_items.append({
                            "제품코드": selected["제품코드"],
                            "정식제품명": selected["정식제품명"],
                            "검색별칭": selected["별칭(검색어)"],
                            "규격": selected["규격"],
                            "단위": selected["단위"],
                            "수량": int(qty),
                        })
                        st.rerun()
            else:
                st.caption("검색 결과에서 제품을 선택하세요.")

        with st.container(border=True):
            st.markdown(
                '<div class="card-title">3. 발주 목록 <span style="font-size:14px;">(발주서에는 정식 제품명으로 표시됩니다)</span></div>',
                unsafe_allow_html=True,
            )

            if st.session_state.order_items:
                df = pd.DataFrame(st.session_state.order_items)

                df["삭제"] = False
                df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)

                edit_view = df[["삭제", "제품코드", "정식제품명", "규격", "단위", "수량"]].copy()

                edited_df = st.data_editor(
                    edit_view,
                    use_container_width=True,
                    hide_index=True,
                    height=220,
                    disabled=["제품코드", "정식제품명", "규격", "단위"],
                    column_config={
                        "삭제": st.column_config.CheckboxColumn("삭제"),
                        "수량": st.column_config.NumberColumn("수량", min_value=0, step=1),
                    },
                    key="order_items_editor",
                )

                checked = edited_df[edited_df["삭제"] == True].index.tolist()
                qty_changed = False

                for idx, row in edited_df.iterrows():
                    new_qty = safe_int(row["수량"], 0)
                    old_qty = safe_int(st.session_state.order_items[idx]["수량"], 0)

                    if new_qty != old_qty:
                        st.session_state.order_items[idx]["수량"] = new_qty
                        qty_changed = True

                if qty_changed:
                    st.rerun()

                total_count, total_qty = calc_totals(st.session_state.order_items)

                delete_col, _, count_col, qty_col = st.columns([1, 1.4, 1.4, 1.6])

                with delete_col:
                    if st.button("선택 삭제"):
                        if checked:
                            st.session_state.order_items = [
                                item
                                for idx, item in enumerate(st.session_state.order_items)
                                if idx not in checked
                            ]
                            st.rerun()

                with count_col:
                    st.markdown(
                        f'<div style="text-align:right;font-weight:900;padding-top:10px;">총 품목수&nbsp;&nbsp;<span style="font-size:22px;">{fmt_int(total_count)} 건</span></div>',
                        unsafe_allow_html=True,
                    )

                with qty_col:
                    st.markdown(
                        f'<div style="text-align:right;font-weight:900;padding-top:10px;">총 수량&nbsp;&nbsp;<span style="font-size:24px;color:#0052cc;">{fmt_int(total_qty)} EA</span></div>',
                        unsafe_allow_html=True,
                    )

                st.markdown(
                    '<div class="notice">※ 발주서에는 정식 제품명과 수량만 표시됩니다.</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.info("발주 목록이 없습니다.")

        with st.container(border=True):
            st.markdown('<div class="card-title">4. 요청사항 및 발주 저장</div>', unsafe_allow_html=True)

            order_date_value = st.date_input(
                "발주일자",
                value=datetime.now().date(),
                key="order_date",
            )

            request_note = st.text_area(
                "요청사항",
                value=st.session_state.get("loaded_request_note", ""),
                height=72,
                label_visibility="collapsed",
            )

            order_date_text = order_date_value.strftime("%Y-%m-%d")

            b1, b2, b3, b4 = st.columns([1, 1.1, 1.1, 1.1], gap="small")

            with b1:
                if st.button("임시저장", use_container_width=True):
                    draft_id = save_draft(vendor["거래처명"], request_note, st.session_state.order_items)
                    st.success(f"임시저장 완료: {draft_id}")

            with b2:
                with st.popover("임시저장 불러오기", use_container_width=True):
                    if drafts.empty:
                        st.caption("임시저장된 발주서가 없습니다.")
                    else:
                        draft_pick = st.selectbox("불러올 임시저장", drafts["임시ID"].tolist(), key="draft_pick_in_write")
                        if st.button("불러오기", key="load_draft_in_write"):
                            draft_header = drafts[drafts["임시ID"] == draft_pick].iloc[0]
                            rows = draft_items[draft_items["임시ID"] == draft_pick].copy()
                            if not rows.empty:
                                st.session_state.order_items = normalize_order_items(rows)
                                st.session_state.loaded_vendor_name = draft_header["거래처명"]
                                st.session_state.loaded_request_note = draft_header.get("요청사항", "")
                                st.rerun()

            with b3:
                if st.button("발주완료", type="primary", use_container_width=True):
                    order_id = save_order(vendor["거래처명"], request_note, st.session_state.order_items)
                    st.success(f"발주 완료 처리되었습니다: {order_id}")

            with b4:
                path = create_excel(vendor, st.session_state.order_items, request_note, order_date_text)
                with open(path, "rb") as file:
                    st.download_button("엑셀 저장", file, file_name=path.name, use_container_width=True)

            st.caption("ⓘ 임시저장은 언제든지 불러와서 수정 후 발주할 수 있습니다.")

    with right:
        with st.container(border=True):
            render_purchase_preview(vendor, st.session_state.order_items, request_note if "request_note" in locals() else "", order_date_text if "order_date_text" in locals() else datetime.now().strftime("%Y-%m-%d"))


def main():
    inject_css()
    init_state()

    vendors, products, aliases, drafts, draft_items, orders, order_items_saved = load_data()
    page = render_sidebar()

    if page == "발주 작성":
        page_write(vendors, products, aliases, drafts, draft_items, orders, order_items_saved)
    elif page == "임시저장 목록":
        page_drafts(drafts, draft_items)
    elif page == "발주서 목록":
        page_orders(vendors, orders, order_items_saved)
    elif page == "최근 발주 내역":
        page_recent_orders(orders)
    elif page == "거래처 관리":
        page_vendor_manage(vendors)
    elif page == "제품 관리":
        page_product_manage(products)
    elif page == "별칭 관리":
        page_alias_manage(vendors, products, aliases)
    else:
        page_placeholder(page)


if __name__ == "__main__":
    main()
